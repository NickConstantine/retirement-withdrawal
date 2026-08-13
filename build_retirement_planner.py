"""Build the Retirement Withdrawal Strategies planner workbook.

Single source of truth for the spreadsheet: edit this file, then run

    python build_retirement_planner.py
    powershell -ExecutionPolicy Bypass -File recalc_excel.ps1
    python validate_model.py
    python patch_metadata.py

Shared constants, historical data and default inputs live in `model_spec.py`
so that `validate_model.py` can recompute the whole model independently and
compare its answers with the values Excel calculated.
"""
import os
import random

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import model_spec as spec

# ---- Style helpers -------------------------------------------------
ARIAL = "Arial"
BLACK = "000000"
GREEN = "008000"
WHITE = "FFFFFF"
BLUE = "0000FF"
MS_BLUE = "0078D4"
LIGHT_GREY = "F3F2F1"
YELLOW = "FFFF00"
PALE_BLUE = "DEECF9"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
AMBER_FILL = "FFEB9C"
AMBER_FONT = "9C6500"

UNLOCK = Protection(locked=False)

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%'
PCT2 = '0.00%'
NUM = '0'
NUM1 = '0.0'

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom_only = Border(bottom=Side(style="thin", color="808080"))
box = Border(left=Side(style="thin", color=MS_BLUE), right=Side(style="thin", color=MS_BLUE),
             top=Side(style="thin", color=MS_BLUE), bottom=Side(style="thin", color=MS_BLUE))


def font(bold=False, color=BLACK, size=11, italic=False):
    return Font(name=ARIAL, bold=bold, color=color, size=size, italic=italic)


def header_cell(c, text):
    c.value = text
    c.font = font(bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=MS_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all


def title(ws_, cell, text, size=16):
    ws_[cell] = text
    ws_[cell].font = font(bold=True, color="242424", size=size)


def band(ws_, row, col_from, col_to, text):
    ws_.cell(row, col_from, text).font = font(bold=True, color=WHITE)
    for c in range(col_from, col_to + 1):
        ws_.cell(row, c).fill = PatternFill("solid", start_color=MS_BLUE)


red_rule = CellIsRule(operator="lessThanOrEqual", formula=["0"],
                      fill=PatternFill("solid", start_color=RED_FILL),
                      font=Font(name=ARIAL, color=RED_FONT))

wb = Workbook()

# ===================================================================
# SHEET: Inputs & Summary  (skeleton + input registry)
# ===================================================================
ws = wb.active
ws.title = "Inputs & Summary"
ws.sheet_view.showGridLines = False
for col, width in (("A", 2), ("B", 54), ("C", 16), ("D", 3), ("E", 56), ("F", 16),
                   ("G", 3), ("H", 8), ("I", 12), ("J", 12), ("K", 12), ("L", 12)):
    ws.column_dimensions[col].width = width

title(ws, "B2", "Retirement Withdrawal Strategies")
ws["B3"] = ("Compare the classic 4% rule with a dynamic strategy (life expectancy + guardrails + "
            "shock absorber) — after tax, fees, healthcare, long-term care and RMDs.  "
            "Yellow = edit these · Black = calculated · Green = from another tab.")
ws["B3"].font = font(italic=True, color="595959", size=9)
ws.merge_cells("B3:F3")

ws.merge_cells("B4:F4")
ws.row_dimensions[4].height = 62
vcell = ws["B4"]
vcell.fill = PatternFill("solid", start_color=PALE_BLUE)
vcell.border = box
vcell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
vcell.font = font(bold=True, color="1F3864", size=11)

_dv = {}


def _validation(kind):
    if kind in _dv:
        return _dv[kind]
    if kind == "age":
        d = DataValidation(type="whole", operator="between", formula1="40", formula2="110")
        d.error = "Enter a whole age between 40 and 110."
    elif kind == "ssage":
        d = DataValidation(type="whole", operator="between", formula1="62", formula2="70")
        d.error = "Social Security start age must be a whole number from 62 to 70."
    elif kind == "rmdage":
        d = DataValidation(type="whole", operator="between", formula1="70", formula2="80")
        d.error = "RMD start age must be a whole number from 70 to 80."
    elif kind == "rate":
        d = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.25")
        d.error = "Enter a rate between 0% and 25%."
    elif kind == "fee":
        d = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.05")
        d.error = "Enter an annual fee between 0% and 5%."
    elif kind == "share":
        d = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
        d.error = "Enter a percentage between 0% and 100%."
    elif kind == "tax":
        d = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.6")
        d.error = "Enter an effective tax rate between 0% and 60%."
    elif kind == "drift":
        d = DataValidation(type="decimal", operator="between", formula1="-0.03", formula2="0.03")
        d.error = "Enter a real spending drift between -3% and +3% per year."
    elif kind == "money":
        d = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
        d.error = "Enter a non-negative dollar amount."
    elif kind == "years":
        d = DataValidation(type="whole", operator="between", formula1="0", formula2="20")
        d.error = "Enter a whole number of years between 0 and 20."
    elif kind == "gap":
        d = DataValidation(type="whole", operator="between", formula1="-20", formula2="20")
        d.error = "Enter a whole number of years between -20 and 20."
    elif kind == "mult":
        d = DataValidation(type="decimal", operator="between", formula1="0.5", formula2="5")
        d.error = "Enter a multiple between 0.5 and 5.0."
    elif kind == "le":
        d = DataValidation(type="decimal", operator="between", formula1="0", formula2="60")
        d.error = "Enter remaining years between 0 and 60."
    else:
        raise KeyError(kind)
    d.allow_blank = False
    ws.add_data_validation(d)
    _dv[kind] = d
    return d


class InputBlock:
    """Assigns rows to inputs as they are declared so no row number is hard-coded."""

    def __init__(self, sheet, first_row, label_col=2, value_col=3, width=2):
        self.ws = sheet
        self.row = first_row
        self.lc = label_col
        self.vc = value_col
        self.width = width
        self.first = first_row

    def section(self, text):
        if self.row > self.first:
            self.row += 1
        band(self.ws, self.row, self.lc, self.lc + self.width - 1, text)
        self.row += 1

    def add(self, label, value, fmt, kind, note=None, choices=None):
        c_label = self.ws.cell(self.row, self.lc, label)
        c_label.font = font()
        c_label.alignment = Alignment(wrap_text=True, vertical="center")
        c = self.ws.cell(self.row, self.vc, value)
        c.font = font(color=BLACK)
        c.number_format = fmt
        c.fill = PatternFill("solid", start_color=YELLOW)
        c.border = bottom_only
        c.alignment = Alignment(horizontal="right")
        c.protection = UNLOCK
        addr = f"{get_column_letter(self.vc)}{self.row}"
        if choices:
            d = DataValidation(type="list", formula1='"' + ",".join(choices) + '"',
                               allow_blank=False)
            d.error = "Choose one of: " + ", ".join(choices)
            self.ws.add_data_validation(d)
            d.add(addr)
        elif kind:
            _validation(kind).add(addr)
        if note:
            cm = Comment(note, "Planner")
            cm.width, cm.height = 300, 120
            c_label.comment = cm
        self.row += 1
        return f"${get_column_letter(self.vc)}${self.row - 1}"


IB = InputBlock(ws, 6)

IB.section("YOU & YOUR HOUSEHOLD")
start_age = IB.add("Retirement start age (or your current age, if already retired)",
                   spec.DEFAULTS["start_age"], NUM, "age",
                   "Year 1 of the tables is this age. If you are already retired, set this to "
                   "your current age and the portfolio below to your current balance.")
plan_age = IB.add("Plan-to age (planning horizon)", spec.DEFAULTS["plan_age"], NUM, "age",
                  "The model reports your balance AT this age. Years modelled = plan-to age "
                  "minus start age.")
household = IB.add("Household", spec.DEFAULTS["household"], "General", None,
                   "Couple uses last-survivor (joint) life expectancy, which is materially "
                   "longer than a single life and therefore a safer divisor.",
                   choices=["Single", "Couple"])
spouse_gap = IB.add("Spouse is younger by (years; negative = older)",
                    spec.DEFAULTS["spouse_gap"], NUM, "gap",
                    "Only used when Household = Couple.")

IB.section("YOUR ACCOUNTS (the starting portfolio is the sum of these)")
taxable0 = IB.add("Taxable / brokerage balance", spec.DEFAULTS["taxable0"], CUR, "money",
                  "Spent first. Only the GAIN is taxed, at the capital-gains rate, and its "
                  "dividends are taxed every year.")
basis_share = IB.add("Cost basis in the taxable account (% of its value)",
                     spec.DEFAULTS["basis_share"], PCT, "share",
                     "What you originally paid, as a share of today's value. 100% means no "
                     "embedded gain; 40% means most of the account is unrealised profit and "
                     "selling it is expensive.")
traditional0 = IB.add("Traditional 401(k) / IRA balance (tax-deferred)",
                      spec.DEFAULTS["traditional0"], CUR, "money",
                      "Spent second. Every dollar is taxed as ordinary income, and this is the "
                      "balance RMDs are calculated from.")
roth0 = IB.add("Roth balance (tax-free)", spec.DEFAULTS["roth0"], CUR, "money",
               "Spent last, so it compounds untouched for as long as possible. Withdrawals are "
               "tax-free and it is not subject to RMDs.")

IB.section("PORTFOLIO & MARKETS")
exp_ret = IB.add("Expected annual return (nominal, BEFORE fees)",
                 spec.DEFAULTS["exp_ret"], PCT, "rate",
                 "Gross return assumption. The fee below is deducted from it every year.")
fee = IB.add("Annual fees (fund expenses + any advisory fee)",
             spec.DEFAULTS["fee"], PCT2, "fee",
             "Deducted from the portfolio every year. Half a percent compounded over 40 years "
             "is a large cumulative drag.")
infl = IB.add("Inflation rate (long-run average)", spec.DEFAULTS["infl"], PCT, "rate")
scenario = IB.add("Market scenario (drives the return AND inflation columns)",
                  spec.DEFAULTS["scenario"], "General", None,
                  "Steady uses your expected return and inflation every year. Early crash and "
                  "Stagflation use the editable table to the right.",
                  choices=["Steady", "Early crash", "Stagflation"])

IB.section("SPENDING (today's dollars, AFTER tax)")
essential = IB.add("Essential annual spending (today's $, after tax)",
                   spec.DEFAULTS["essential"], CUR, "money",
                   "Your must-have spending. This is an AFTER-TAX number, which is why the "
                   "model computes after-tax income to compare against it.")
healthcare = IB.add("Extra healthcare premiums before age 65 (today's $/yr)",
                    spec.DEFAULTS["healthcare"], CUR, "money",
                    "Pre-Medicare coverage, added on top of essential spending for every year "
                    "you are under 65.")
drift = IB.add("Real spending drift per year (spending smile)",
               spec.DEFAULTS["drift"], PCT2, "drift",
               "Research (Blanchett) finds real spending drifts down roughly 1%/yr through "
               "mid-retirement. 0% holds essentials flat in real terms and is the "
               "conservative choice.")

IB.section("STRATEGY SETTINGS")
rate4 = IB.add("4% rule starting rate", spec.DEFAULTS["rate4"], PCT, "rate")
floor = IB.add("Guardrail floor (minimum withdrawal rate)", spec.DEFAULTS["floor"], PCT, "rate")
ceiling = IB.add("Guardrail ceiling (maximum withdrawal rate)",
                 spec.DEFAULTS["ceiling"], PCT, "rate",
                 "Only used when the ceiling basis below is set to Fixed %. Watch the "
                 "diagnostics: a flat ceiling is age-blind, so as your remaining years shrink "
                 "it progressively overrides the life-expectancy strategy instead of just "
                 "trimming excess.")
ceiling_basis = IB.add("Guardrail ceiling basis", spec.DEFAULTS["ceiling_basis"], "General", None,
                       "Fixed %: one flat cap at every age, exactly as the source article "
                       "describes. Age-graduated: the cap rises with age, set at a multiple of "
                       "the IRS Uniform Lifetime withdrawal rate (the same published table used "
                       "for RMDs). Graduated lets the actuarial rule keep working into your 80s "
                       "while still blocking a genuinely extreme withdrawal at 95.",
                       choices=["Age-graduated", "Fixed %"])
ceiling_mult = IB.add("Age-graduated ceiling — multiple of the IRS rate",
                      spec.DEFAULTS["ceiling_mult"], NUM1, "mult",
                      "The cap is this multiple of the IRS Uniform Lifetime withdrawal rate for "
                      "your age. At 2.0 the cap first bites around age 83. Lower values cap "
                      "sooner and leave a larger estate; above about 2.25 the cap never binds at "
                      "all within a normal horizon.")
shock = IB.add("Shock absorber — max spending change per year",
               spec.DEFAULTS["shock"], PCT, "rate")
shock_basis = IB.add("Shock absorber applies to", spec.DEFAULTS["shock_basis"], "General", None,
                     "Real: the band is applied after adjusting last year's withdrawal for "
                     "inflation, so a 5% cut is a 5% cut in buying power. Nominal: the band is "
                     "applied to raw dollars, which at 3% inflation is really +1.9% / -7.8%.",
                     choices=["Real", "Nominal"])
current_4pct = IB.add("Current 4% rule withdrawal if already retired (today's $/yr; 0 if not)",
                      spec.DEFAULTS["current_4pct"], CUR, "money")

IB.section("TAXES & REQUIRED DISTRIBUTIONS")
tax_early = IB.add("Effective tax rate — before Social Security starts",
                   spec.DEFAULTS["tax_early"], PCT, "tax",
                   "In early retirement your only taxable income is what you withdraw, so most "
                   "of it lands in the lowest brackets. Using one flat lifetime rate is the most "
                   "common way a projection understates early-retirement spending.")
tax_ss = IB.add("Effective tax rate — Social Security until RMDs",
                spec.DEFAULTS["tax_ss"], PCT, "tax",
                "Once Social Security starts, taxable income rises and up to 85% of the benefit "
                "itself becomes taxable.")
tax_rmd = IB.add("Effective tax rate — once RMDs begin",
                 spec.DEFAULTS["tax_rmd"], PCT, "tax",
                 "Required distributions stack on top of Social Security, usually making this "
                 "the highest-tax phase of retirement.")
cg_rate = IB.add("Long-term capital gains effective rate",
                 spec.DEFAULTS["cg_rate"], PCT, "tax",
                 "Applied to the GAIN portion of taxable-account sales and to its dividends. "
                 "0/15/20% federally, plus 3.8% NIIT at higher incomes, plus any state tax.")
div_yield = IB.add("Taxable account dividend yield (taxed every year)",
                   spec.DEFAULTS["div_yield"], PCT, "rate",
                   "Dividends and interest are taxed as they are paid, whether or not you spend "
                   "them. Ignoring this overstates a taxable account's compounding.")
rmd_age = IB.add("RMD start age", spec.DEFAULTS["rmd_age"], NUM, "rmdage",
                 "SECURE 2.0: age 73, or 75 if born 1960 or later. RMDs are modelled as a tax "
                 "drag rather than forced spending — any excess above what you wanted to spend "
                 "is taken from the traditional account, taxed, and reinvested in the taxable "
                 "account.")

IB.section("GUARANTEED INCOME — SOCIAL SECURITY")
full_ss = IB.add("Full Social Security benefit at FRA 67 (today's $/yr, before tax)",
                 spec.DEFAULTS["full_ss"], CUR, "money")
ss_age = IB.add("Social Security start age (62-70)", spec.DEFAULTS["ss_age"], NUM, "ssage")

IB.section("LONG-TERM CARE")
ltc_on = IB.add("Include a long-term care event?", spec.DEFAULTS["ltc_on"], "General", None,
                "Roughly half of 65-year-olds will need some paid long-term care. Leaving it "
                "out is the most common way a retirement plan looks safer than it is.",
                choices=["Yes", "No"])
ltc_cost = IB.add("Long-term care cost (today's $/yr)", spec.DEFAULTS["ltc_cost"], CUR, "money")
ltc_years = IB.add("Long-term care duration (years)", spec.DEFAULTS["ltc_years"], NUM, "years")
ltc_age = IB.add("Long-term care starting age", spec.DEFAULTS["ltc_age"], NUM, "age")

IB.section("LIFE EXPECTANCY ANCHORS (from the SSA calculator)")
cur_age_le = IB.add("Current age (for life expectancy)", spec.DEFAULTS["cur_age_le"], NUM, "age")
le_cur = IB.add("Remaining life expectancy at that current age",
                spec.DEFAULTS["le_cur"], NUM1, "le")
le62 = IB.add("Remaining life expectancy at age 62", spec.DEFAULTS["le62"], NUM1, "le")
le67 = IB.add("Remaining life expectancy at age 67", spec.DEFAULTS["le67"], NUM1, "le")
le70 = IB.add("Remaining life expectancy at age 70", spec.DEFAULTS["le70"], NUM1, "le")

S = "'Inputs & Summary'!"


def q(ref):
    return S + ref


bare_plan = plan_age.replace("$", "")
bare_start = start_age.replace("$", "")
dv_plan = DataValidation(
    type="custom",
    formula1=f"=AND({bare_plan}>={bare_start},{bare_plan}<=110,{bare_plan}=INT({bare_plan}))",
    allow_blank=False)
dv_plan.error = "Plan-to age must be a whole number from the retirement start age through 110."
ws.add_data_validation(dv_plan)
dv_plan.add(bare_plan)

# Years actually lived through: ages start_age .. plan_age-1. The balance reported
# "at plan-to age" is the ENDING balance of the last of those years.
HORIZON = f"MIN(MAX({q(plan_age)}-{q(start_age)},1),{spec.TABLE_ROWS})"

# Everything from column D rightward starts one row lower than the input block, so
# row 5 reads as a blank separator under the verdict box on both sides of the sheet.
# Columns A-C already have a natural gap there (the input block starts at row 6).
PANEL_TOP = 6

# ---- SCENARIO DETAIL table (per-year return AND inflation) ----
ws.merge_cells(f"H{PANEL_TOP}:L{PANEL_TOP}")
band(ws, PANEL_TOP, 8, 12, "SCENARIO DETAIL — annual return and inflation (edit yellow cells)")
header_cell(ws.cell(PANEL_TOP + 1, 8), "Year")
header_cell(ws.cell(PANEL_TOP + 1, 9), "Early crash return")
header_cell(ws.cell(PANEL_TOP + 1, 10), "Early crash inflation")
header_cell(ws.cell(PANEL_TOP + 1, 11), "Stagflation return")
header_cell(ws.cell(PANEL_TOP + 1, 12), "Stagflation inflation")
ws.row_dimensions[PANEL_TOP + 1].height = 30

dv_scenret = DataValidation(type="decimal", operator="between", formula1="-0.6", formula2="0.6",
                            allow_blank=False)
dv_scenret.error = "Enter an annual rate between -60% and 60%."
ws.add_data_validation(dv_scenret)

SCEN_FIRST = PANEL_TOP + 2
for j in range(spec.SCENARIO_YEARS):
    r_ = SCEN_FIRST + j
    yv = j + 1
    yc = ws.cell(r_, 8, yv)
    yc.font = font()
    yc.alignment = Alignment(horizontal="center")
    # Early crash: bad returns up front, inflation stays at the base assumption
    if yv in spec.CRASH_RETURNS:
        c_ = ws.cell(r_, 9, spec.CRASH_RETURNS[yv])
        c_.fill = PatternFill("solid", start_color=YELLOW)
        c_.protection = UNLOCK
        c_.font = font()
        dv_scenret.add(f"I{r_}")
    else:
        c_ = ws.cell(r_, 9, f"={q(exp_ret)}")
        c_.font = font(color="595959")
    c_.number_format = PCT
    ci = ws.cell(r_, 10, f"={q(infl)}")
    ci.font = font(color="595959")
    ci.number_format = PCT
    # Stagflation: weak returns AND high inflation - that combination is what
    # actually destroyed the 1966 retiree cohort.
    if yv <= spec.STAGFLATION_YEARS:
        c2 = ws.cell(r_, 11, spec.STAGFLATION_RETURN)
        c2.fill = PatternFill("solid", start_color=YELLOW)
        c2.protection = UNLOCK
        c2.font = font()
        dv_scenret.add(f"K{r_}")
        c3 = ws.cell(r_, 12, spec.STAGFLATION_INFLATION)
        c3.fill = PatternFill("solid", start_color=YELLOW)
        c3.protection = UNLOCK
        c3.font = font()
        dv_scenret.add(f"L{r_}")
    else:
        c2 = ws.cell(r_, 11, f"={q(exp_ret)}")
        c2.font = font(color="595959")
        c3 = ws.cell(r_, 12, f"={q(infl)}")
        c3.font = font(color="595959")
    c2.number_format = PCT
    c3.number_format = PCT
SCEN_LAST = SCEN_FIRST + spec.SCENARIO_YEARS - 1
SCEN_TBL = f"{S}$H${SCEN_FIRST}:$L${SCEN_LAST}"
ws.cell(SCEN_LAST + 1, 8,
        "Years beyond 10 use your expected return and inflation. Steady uses them in every year."
        ).font = font(size=8, italic=True, color="808080")
ws.merge_cells(start_row=SCEN_LAST + 1, start_column=8, end_row=SCEN_LAST + 1, end_column=12)


def scen_return(age_row_col):
    """Formula fragment: this year's gross return for the chosen scenario."""
    return (f'IF({q(scenario)}="Steady",{q(exp_ret)},'
            f'IF({q(scenario)}="Early crash",'
            f'IFERROR(VLOOKUP({age_row_col},{SCEN_TBL},2,FALSE),{q(exp_ret)}),'
            f'IFERROR(VLOOKUP({age_row_col},{SCEN_TBL},4,FALSE),{q(exp_ret)})))')


def scen_inflation(age_row_col):
    """Formula fragment: this year's inflation for the chosen scenario."""
    return (f'IF({q(scenario)}="Steady",{q(infl)},'
            f'IF({q(scenario)}="Early crash",'
            f'IFERROR(VLOOKUP({age_row_col},{SCEN_TBL},3,FALSE),{q(infl)}),'
            f'IFERROR(VLOOKUP({age_row_col},{SCEN_TBL},5,FALSE),{q(infl)})))')


# ===================================================================
# SHEET: Reference Tables
# ===================================================================
ref = wb.create_sheet("Reference Tables")
ref.sheet_view.showGridLines = False
title(ref, "A1", "Reference Tables", 14)
ref["A2"] = ("Published tables the model looks up. These are facts, not assumptions — change "
             "them only if the underlying rules change.")
ref["A2"].font = font(italic=True, color="595959", size=9)
for col, width in (("A", 12), ("B", 16), ("C", 4), ("D", 12), ("E", 16)):
    ref.column_dimensions[col].width = width

band(ref, 4, 1, 2, "SOCIAL SECURITY CLAIMING FACTORS")
header_cell(ref.cell(5, 1), "Start age")
header_cell(ref.cell(5, 2), "% of full benefit")
SS_FIRST = 6
for i, (age_, fac_) in enumerate(spec.SS_FACTORS):
    r_ = SS_FIRST + i
    a_ = ref.cell(r_, 1, age_)
    a_.font = font()
    a_.alignment = Alignment(horizontal="center")
    a_.number_format = NUM
    f_ = ref.cell(r_, 2, fac_)
    f_.font = font(color=BLUE)
    f_.number_format = "0.0%"
    f_.alignment = Alignment(horizontal="center")
SS_LAST = SS_FIRST + len(spec.SS_FACTORS) - 1
ref.cell(SS_LAST + 2, 1,
         "Source: SSA, born 1960 or later (full retirement age 67). "
         "ssa.gov/benefits/retirement/planner/1960.html").font = font(size=8, italic=True,
                                                                      color="808080")
ref.merge_cells(start_row=SS_LAST + 2, start_column=1, end_row=SS_LAST + 2, end_column=2)

band(ref, 4, 4, 5, "IRS UNIFORM LIFETIME TABLE (RMD)")
header_cell(ref.cell(5, 4), "Age")
header_cell(ref.cell(5, 5), "Divisor")
RMD_FIRST = 6
for i, (age_, div_) in enumerate(spec.RMD_DIVISORS):
    r_ = RMD_FIRST + i
    a_ = ref.cell(r_, 4, age_)
    a_.font = font()
    a_.alignment = Alignment(horizontal="center")
    a_.number_format = NUM
    d_ = ref.cell(r_, 5, div_)
    d_.font = font(color=BLUE)
    d_.number_format = NUM1
    d_.alignment = Alignment(horizontal="center")
RMD_LAST = RMD_FIRST + len(spec.RMD_DIVISORS) - 1
ref.cell(RMD_LAST + 2, 4, "Source: IRS Publication 590-B, Table III (2022 onward).").font = font(
    size=8, italic=True, color="808080")
ref.merge_cells(start_row=RMD_LAST + 2, start_column=4, end_row=RMD_LAST + 2, end_column=5)

SS_TBL = f"'Reference Tables'!$A${SS_FIRST}:$B${SS_LAST}"
RMD_TBL = f"'Reference Tables'!$D${RMD_FIRST}:$E${RMD_LAST}"

SS_FACTOR = f"VLOOKUP({q(ss_age)},{SS_TBL},2,FALSE)"

# ---------------------------------------------------------------------------
# Account model
#
# Three balances are tracked: taxable, traditional (tax-deferred) and Roth.
# A gross withdrawal is sourced taxable-first, then traditional, then Roth --
# the conventional order, which lets the sheltered accounts compound longest.
#
#   * taxable      only the GAIN is taxed, at the capital-gains rate, and the
#                  account's dividends are taxed every year whether spent or not
#   * traditional  every dollar is ordinary income, taxed at the rate for the
#                  current retirement phase, and RMDs are calculated from it
#   * Roth         tax-free, no RMDs
#
# Because a gross withdrawal is never allowed to exceed the total balance, the
# Roth slice needs no MIN() guard: whatever the first two cannot cover fits.
# ---------------------------------------------------------------------------


def ord_rate(age_cell):
    """Effective ordinary tax rate for the retirement phase this age falls in."""
    return (f"IF({age_cell}>={q(rmd_age)},{q(tax_rmd)},"
            f"IF({age_cell}>={q(ss_age)},{q(tax_ss)},{q(tax_early)}))")


def from_taxable(w, tx):
    return f"MIN({w},{tx})"


def from_traditional(w, tx, td):
    return f"MIN(MAX({w}-{tx},0),{td})"


def from_roth(w, tx, td):
    return f"MAX({w}-{tx}-MAX(MIN(MAX({w}-{tx},0),{td}),0),0)"


def gain_fraction(tx, basis):
    """Share of a taxable-account sale that is taxable gain."""
    return f"IF({tx}>0,MAX(1-{basis}/{tx},0),0)"


def rmd_from(age_cell, td):
    return (f"IF({age_cell}>={q(rmd_age)},"
            f"{td}/VLOOKUP(MIN(MAX({age_cell},72),120),{RMD_TBL},2,TRUE),0)")


def ss_gross(age_cell, index_cell):
    """Nominal Social Security in a given year (grown by the actual price index)."""
    return f"IF({age_cell}>={q(ss_age)},{q(full_ss)}*{SS_FACTOR}*{index_cell},0)"


def need_today(year_cell, age_cell):
    """Total spending requirement for the year, in today's dollars, after tax."""
    return (f"{q(essential)}*(1+{q(drift)})^({year_cell}-1)"
            f"+IF({age_cell}<65,{q(healthcare)},0)"
            f"+IF(AND({q(ltc_on)}=\"Yes\",{age_cell}>={q(ltc_age)},"
            f"{age_cell}<{q(ltc_age)}+{q(ltc_years)}),{q(ltc_cost)},0)")


GROWTH = f"((1+{{r}})*(1-{q(fee)}))"


# ===================================================================
# SHEET: Life Expectancy
#   Anchors are sorted before use so an out-of-order "current age" cannot
#   silently corrupt the interpolation. Implied one-year survival is derived
#   from the table (p(x) = e(x)/(1+e(x+1))), which gives an exact survival
#   curve and therefore a proper last-survivor life expectancy for couples.
# ===================================================================
le = wb.create_sheet("Life Expectancy")
le.sheet_view.showGridLines = False
title(le, "A1", "Remaining Life Expectancy — calculated", 14)
le["A2"] = ("Interpolated from the four anchors on the Inputs tab (your current age, 62, 67, 70). "
            "Ages past 70 scale from the age-70 value. The survival columns turn that table into "
            "a year-by-year survival curve, which is what makes joint (couple) life expectancy "
            "possible.")
le["A2"].font = font(italic=True, color="595959", size=9)
le.merge_cells("A2:F2")
le["A3"] = "Anchor source: SSA Life Expectancy Calculator (ssa.gov/OACT/population/longevity.html)."
le["A3"].font = font(italic=True, color="808080", size=8)
for col, width in (("A", 8), ("B", 15), ("C", 26), ("E", 11), ("F", 13), ("H", 13),
                   ("I", 14), ("J", 15), ("K", 15), ("L", 17)):
    le.column_dimensions[col].width = width

# ---- anchor table (E/F), always sorted ascending by age ----
header_cell(le.cell(5, 5), "Anchor age")
header_cell(le.cell(5, 6), "Anchor years")
ANCHOR_FIRST = 6
anchor_pairs = [(f"={q(cur_age_le)}", f"={q(le_cur)}"),
                ("=62", f"={q(le62)}"),
                ("=67", f"={q(le67)}"),
                ("=70", f"={q(le70)}")]
for age_, ratio in spec.LE_TAIL_ANCHORS:
    anchor_pairs.append((f"={age_}", f"={q(le70)}*{ratio}"))
N_ANCHOR = len(anchor_pairs)
ANCHOR_LAST = ANCHOR_FIRST + N_ANCHOR - 1

# Raw (possibly unsorted) anchors are written to a helper block, then ranked into
# ascending order. MATCH(...,1) requires ascending data; the previous build fed it
# an unsorted list whenever the "current age" anchor was above 62.
RAW_FIRST = ANCHOR_FIRST
RAW_COL, RAWL_COL = 14, 15          # N, O  (hidden helper columns)
le.cell(5, RAW_COL, "raw age").font = font(size=8, color="808080")
le.cell(5, RAWL_COL, "raw years").font = font(size=8, color="808080")
for i, (a_formula, l_formula) in enumerate(anchor_pairs):
    r_ = RAW_FIRST + i
    ca = le.cell(r_, RAW_COL, a_formula)
    ca.font = font(size=8, color="808080")
    ca.number_format = NUM
    cl = le.cell(r_, RAWL_COL, l_formula)
    cl.font = font(size=8, color="808080")
    cl.number_format = NUM1
RAW_LAST = RAW_FIRST + N_ANCHOR - 1
RAW_AGES = f"$N${RAW_FIRST}:$N${RAW_LAST}"
RAW_LES = f"$O${RAW_FIRST}:$O${RAW_LAST}"

for i in range(N_ANCHOR):
    r_ = ANCHOR_FIRST + i
    k = i + 1
    # k-th smallest age; ties are broken by nudging duplicates upward so that
    # MATCH still sees a strictly ascending column.
    ca = le.cell(r_, 5, f"=SMALL({RAW_AGES},{k})+{i}*0.0001")
    ca.font = font(color="808080")
    ca.number_format = NUM
    # the life expectancy that belongs to that age (INDEX/MATCH avoids needing
    # an array-entered formula, which older Excel versions require for MIN(IF(...)))
    cl = le.cell(r_, 6,
                 f"=INDEX({RAW_LES},MATCH(SMALL({RAW_AGES},{k}),{RAW_AGES},0))")
    cl.font = font(color="808080")
    cl.number_format = NUM1
ANCHOR_AGES = f"$E${ANCHOR_FIRST}:$E${ANCHOR_LAST}"
ANCHOR_LES = f"$F${ANCHOR_FIRST}:$F${ANCHOR_LAST}"
le.cell(ANCHOR_LAST + 1, 5, "Anchors are sorted automatically.").font = font(
    size=8, italic=True, color="808080")
le.merge_cells(start_row=ANCHOR_LAST + 1, start_column=5, end_row=ANCHOR_LAST + 1, end_column=6)

# ---- main table ----
for col, text in ((1, "Age"), (2, "Remaining years (single)"), (3, "Basis"),
                  (8, "1-yr survival"), (9, "Survival from 35"), (10, "Spouse survival"),
                  (11, "Joint remaining years"), (12, "Planning years (used)"),
                  (13, "Guardrail ceiling (used)")):
    header_cell(le.cell(5, col), text)
le.row_dimensions[5].height = 30

LE_FIRST = 6
AGES = list(range(spec.LE_FIRST_AGE, spec.LE_LAST_AGE + 1))
LE_LAST = LE_FIRST + len(AGES) - 1
AGE_RANGE = f"$A${LE_FIRST}:$A${LE_LAST}"
SURV_RANGE = f"$I${LE_FIRST}:$I${LE_LAST}"
SPSURV_RANGE = f"$J${LE_FIRST}:$J${LE_LAST}"

for i, age_ in enumerate(AGES):
    r_ = LE_FIRST + i
    ac = le.cell(r_, 1, age_)
    ac.font = font()
    ac.alignment = Alignment(horizontal="center")
    m = f"IFERROR(MATCH(A{r_},{ANCHOR_AGES},1),1)"
    al = f"INDEX({ANCHOR_AGES},{m})"
    au = f"INDEX({ANCHOR_AGES},MIN({m}+1,{N_ANCHOR}))"
    ll = f"INDEX({ANCHOR_LES},{m})"
    lu = f"INDEX({ANCHOR_LES},MIN({m}+1,{N_ANCHOR}))"
    b = le.cell(r_, 2,
                f"=MAX(ROUND({ll}+IF({au}={al},0,(A{r_}-{al})/({au}-{al})*({lu}-{ll})),1),0.1)")
    b.font = font()
    b.number_format = NUM1
    le.cell(r_, 3,
            f'=IF(A{r_}<=MAX({ANCHOR_AGES}),IF(A{r_}>=MIN({ANCHOR_AGES}),'
            f'"Interpolated","Extrapolated below anchors"),"Extrapolated above anchors")'
            ).font = font(size=9, color="808080")
    # implied one-year survival: e(x) = p(x) * (1 + e(x+1))
    nxt = f"B{r_ + 1}" if age_ < spec.LE_LAST_AGE else f"B{r_}*0.84"
    ps = le.cell(r_, 8, f"=MIN(MAX(B{r_}/(1+{nxt}),0),1)")
    ps.font = font(color="808080")
    ps.number_format = "0.0000"
    # cumulative survival from the first age in the table
    sv = le.cell(r_, 9, "=1" if i == 0 else f"=I{r_ - 1}*H{r_ - 1}")
    sv.font = font(color="808080")
    sv.number_format = "0.000000"
    # the spouse's cumulative survival at the same point in calendar time
    sp = le.cell(r_, 10,
                 f"=IFERROR(INDEX({SURV_RANGE},MATCH(A{r_}-{q(spouse_gap)},{AGE_RANGE},0)),"
                 f"IF(A{r_}-{q(spouse_gap)}<{spec.LE_FIRST_AGE},1,0))")
    sp.font = font(color="808080")
    sp.number_format = "0.000000"
    # last-survivor remaining years: sum over future years of P(at least one alive)
    jl = le.cell(r_, 11,
                 f"=IFERROR(SUMPRODUCT(({AGE_RANGE}>A{r_})*"
                 f"(1-(1-{SURV_RANGE}/I{r_})*(1-{SPSURV_RANGE}/J{r_}))),B{r_})")
    jl.font = font(color="808080")
    jl.number_format = NUM1
    pl = le.cell(r_, 12, f'=MAX(IF({q(household)}="Couple",K{r_},B{r_}),0.5)')
    pl.font = font(bold=True)
    pl.number_format = NUM1
    # Age-graduated ceiling = multiple x the IRS Uniform Lifetime withdrawal rate.
    # The published table only covers 72-120, so the age is clamped into that range.
    cg = le.cell(r_, 13,
                 f'=IF({q(ceiling_basis)}="Fixed %",{q(ceiling)},'
                 f'{q(ceiling_mult)}/VLOOKUP(MIN(MAX(A{r_},72),120),{RMD_TBL},2,TRUE))')
    cg.font = font()
    cg.number_format = PCT2
    # helper: the age itself when the ceiling caps the raw actuarial rate, else a
    # sentinel, so the results panel can report the first such age with a plain MIN.
    bd = le.cell(r_, 16,
                 f"=IF(AND(A{r_}>={q(start_age)},A{r_}<{q(plan_age)},1/L{r_}>M{r_}+0.000001),"
                 f"A{r_},9999)")
    bd.font = font(size=8, color="808080")
    bd.number_format = NUM

for col in ("N", "O", "P"):
    le.column_dimensions[col].hidden = True
le.column_dimensions["M"].width = 18

LE_TBL = f"'Life Expectancy'!$A${LE_FIRST}:$M${LE_LAST}"
PLANNING_YEARS_COL = 12
CEILING_COL = 13
BIND_RANGE = f"'Life Expectancy'!$P${LE_FIRST}:$P${LE_LAST}"


def planning_years(age_cell):
    return f"IFERROR(VLOOKUP({age_cell},{LE_TBL},{PLANNING_YEARS_COL},FALSE),1)"


def ceiling_at(age_cell):
    """The guardrail ceiling that applies at this age (flat or age-graduated)."""
    return f"IFERROR(VLOOKUP({age_cell},{LE_TBL},{CEILING_COL},FALSE),{q(ceiling)})"


# ===================================================================
# SHEET: 4% Rule
#   Holds the shared return/inflation path that both strategies use.
# ===================================================================
R0 = 5                                   # first data row on both strategy tabs
RN = R0 + spec.TABLE_ROWS - 1

fr = wb.create_sheet("4% Rule")
fr.sheet_view.showGridLines = False
title(fr, "A1", "4% Rule — Year by Year", 14)
fr["A2"] = ("Year-1 withdrawal is 4% of the starting portfolio; after that the TARGET grows with "
            "actual inflation (it does not reset downward in a year the portfolio cannot fund it). "
            "Returns and inflation come from the Market scenario on the Inputs tab and are shared "
            "with the Dynamic Strategy tab.")
fr["A2"].font = font(italic=True, color="595959", size=9)
fr.merge_cells("A2:T2")

f_cols = [
    ("Year", 6, None),
    ("Age", 6, None),
    ("Annual return", 11, "Gross return before fees for this year, set by the Market scenario."),
    ("Inflation", 10, "This year's inflation, also set by the Market scenario."),
    ("Price index", 10, "Cumulative inflation since year 1. Divide a nominal figure by this to "
                        "get today's buying power."),
    ("Ordinary tax rate", 11, "Your effective rate for this phase: the early rate before Social "
                              "Security, then the Social Security rate, then the RMD rate."),
    ("Taxable beginning", 14, "Brokerage balance at the start of the year. Spent first."),
    ("Traditional beginning", 14, "Tax-deferred balance. Spent second, taxed as ordinary income, "
                                  "and the base for RMDs."),
    ("Roth beginning", 13, "Tax-free balance. Spent last so it compounds longest."),
    ("Taxable cost basis", 13, "What you paid for the taxable holdings. Value above this is the "
                               "gain that gets taxed when you sell."),
    ("Total beginning balance", 15, "The three accounts combined."),
    ("Target withdrawal", 14, "The 4% rule target: 4% of the starting portfolio, then grown by "
                              "actual inflation. It is not reset downward by a lean year."),
    ("Gross withdrawal", 14, "The target, capped at the total balance available."),
    ("From taxable", 12, None),
    ("From traditional", 13, None),
    ("From Roth", 11, None),
    ("RMD", 11, "Required minimum distribution from the traditional balance."),
    ("RMD forced extra", 13, "The part of the RMD you did not want to spend. It is taxed and the "
                             "remainder is reinvested in the taxable account."),
    ("Tax on withdrawal", 13, "Capital-gains tax on the taxable gain plus ordinary tax on the "
                              "traditional dollars. Roth withdrawals are untaxed."),
    ("Dividend tax", 11, "Tax on the taxable account's dividends, paid every year whether or not "
                         "you spend them."),
    ("Social Security", 12, "Nominal benefit this year, zero before your start age."),
    ("Tax on Social Security", 13, "85% of the benefit is treated as taxable."),
    ("After-tax income", 14, "What you can actually spend: withdrawal plus Social Security, both "
                             "after tax."),
    ("After-tax income (today's $)", 16, "The same figure in today's buying power."),
    ("Spending need (today's $)", 16, "Essentials, plus pre-65 healthcare, plus long-term care "
                                      "in the years it applies."),
    ("Surplus / (shortfall)", 15, "After-tax income minus spending need, in today's dollars."),
    ("Taxable ending", 14, None),
    ("Traditional ending", 14, None),
    ("Roth ending", 12, None),
    ("Basis ending", 12, None),
    ("Total ending balance", 15, "Carried into next year."),
    ("Ending balance (today's $)", 16, "Ending balance in today's buying power."),
]
for i, (name, width, note) in enumerate(f_cols, start=1):
    hc = fr.cell(4, i)
    header_cell(hc, name)
    if note:
        cm = Comment(note, "Planner")
        cm.width, cm.height = 280, 110
        hc.comment = cm
    fr.column_dimensions[get_column_letter(i)].width = width
fr.row_dimensions[4].height = 42

for k in range(spec.TABLE_ROWS):
    r_ = R0 + k
    yr = k + 1
    p = r_ - 1                       # previous row
    fr.cell(r_, 1, yr).font = font()
    fr.cell(r_, 2, f"={q(start_age)}+{yr}-1").font = font()
    c = fr.cell(r_, 3, "=" + scen_return(f"A{r_}"))
    c.font = font(color=GREEN)
    c.number_format = PCT
    c = fr.cell(r_, 4, "=" + scen_inflation(f"A{r_}"))
    c.font = font(color=GREEN)
    c.number_format = PCT
    # price index at the START of the year
    c = fr.cell(r_, 5, "=1" if k == 0 else f"=E{p}*(1+D{p})")
    c.font = font(color="595959")
    c.number_format = "0.000"
    c = fr.cell(r_, 6, "=" + ord_rate(f"B{r_}"))
    c.font = font(color="808080")
    c.number_format = PCT
    fr.cell(r_, 7, f"={q(taxable0)}" if k == 0 else f"=MAX(AA{p},0)").font = font()
    fr.cell(r_, 8, f"={q(traditional0)}" if k == 0 else f"=MAX(AB{p},0)").font = font()
    fr.cell(r_, 9, f"={q(roth0)}" if k == 0 else f"=MAX(AC{p},0)").font = font()
    fr.cell(r_, 10, f"={q(taxable0)}*{q(basis_share)}" if k == 0 else f"=MAX(AD{p},0)"
            ).font = font(color="808080")
    fr.cell(r_, 11, f"=G{r_}+H{r_}+I{r_}").font = font()
    # Bengen's target grows with inflation from the previous TARGET, never from a
    # capped actual withdrawal, so one lean year does not permanently cut spending.
    if k == 0:
        fr.cell(r_, 12, f"=IF({q(current_4pct)}>0,{q(current_4pct)},K{r_}*{q(rate4)})")
    else:
        fr.cell(r_, 12, f"=L{p}*(1+D{p})")
    fr.cell(r_, 13, f"=MIN(L{r_},K{r_})").font = font(bold=True)
    fr.cell(r_, 14, "=" + from_taxable(f"M{r_}", f"G{r_}")).font = font()
    fr.cell(r_, 15, "=" + from_traditional(f"M{r_}", f"G{r_}", f"H{r_}")).font = font()
    fr.cell(r_, 16, "=" + from_roth(f"M{r_}", f"G{r_}", f"H{r_}")).font = font()
    fr.cell(r_, 17, "=" + rmd_from(f"B{r_}", f"H{r_}")).font = font(color="808080")
    fr.cell(r_, 18, f"=MAX(MIN(Q{r_},H{r_})-O{r_},0)").font = font(color="808080")
    gain = gain_fraction(f"G{r_}", f"J{r_}")
    fr.cell(r_, 19, f"=N{r_}*{gain}*{q(cg_rate)}+(O{r_}+R{r_})*F{r_}").font = font()
    fr.cell(r_, 20, f"=MAX(G{r_}-N{r_},0)*{q(div_yield)}*{q(cg_rate)}").font = font(color="808080")
    fr.cell(r_, 21, "=" + ss_gross(f"B{r_}", f"E{r_}")).font = font(color="808080")
    fr.cell(r_, 22, f"=U{r_}*{spec.SS_TAXABLE_SHARE}*F{r_}").font = font()
    # the RMD excess is withdrawn but not spent, so its tax is excluded here
    fr.cell(r_, 23,
            f"=M{r_}-(N{r_}*{gain}*{q(cg_rate)}+O{r_}*F{r_})+U{r_}-V{r_}").font = font(bold=True)
    fr.cell(r_, 24, f"=W{r_}/E{r_}").font = font(color="595959")
    fr.cell(r_, 25, "=" + need_today(f"A{r_}", f"B{r_}")).font = font(color="595959")
    fr.cell(r_, 26, f"=X{r_}-Y{r_}").font = font()
    g = f"((1+C{r_})*(1-{q(fee)}))"
    reinvest = f"R{r_}*(1-F{r_})"
    div = f"MAX(G{r_}-N{r_},0)*{q(div_yield)}"
    fr.cell(r_, 27, f"=MAX((G{r_}-N{r_}+{reinvest})*{g}-T{r_},0)").font = font()
    fr.cell(r_, 28, f"=MAX((H{r_}-O{r_}-R{r_})*{g},0)").font = font()
    fr.cell(r_, 29, f"=MAX((I{r_}-P{r_})*{g},0)").font = font()
    # basis falls proportionally with the sale, and rises by reinvested after-tax
    # money and by dividends that have already been taxed
    fr.cell(r_, 30,
            f"=MAX(J{r_}*IF(G{r_}>0,1-N{r_}/G{r_},0)+{reinvest}+{div},0)").font = font(
                color="808080")
    fr.cell(r_, 31, f"=AA{r_}+AB{r_}+AC{r_}").font = font()
    fr.cell(r_, 32, f"=AE{r_}/(E{r_}*(1+D{r_}))").font = font(color="595959")
    for col in range(1, 33):
        cell = fr.cell(r_, col)
        if col >= 7:
            cell.number_format = CUR
        if k % 2 == 1:
            cell.fill = PatternFill("solid", start_color=LIGHT_GREY)

fr.conditional_formatting.add(f"AE{R0}:AE{RN}", red_rule)
fr.conditional_formatting.add(
    f"Z{R0}:Z{RN}",
    FormulaRule(formula=[f"AND(Z{R0}<0,A{R0}<={HORIZON})"],
                fill=PatternFill("solid", start_color=RED_FILL),
                font=Font(name=ARIAL, color=RED_FONT)))

# ===================================================================
# SHEET: Dynamic Strategy
# ===================================================================
dy = wb.create_sheet("Dynamic Strategy")
dy.sheet_view.showGridLines = False
title(dy, "A1", "Dynamic Strategy — Year by Year", 14)
dy["A2"] = ("Withdrawal = balance / remaining planning years, held inside the guardrails, then "
            "limited by the shock absorber, then floored by any RMD tax. Returns and inflation "
            "are linked from the 4% Rule tab so both strategies see identical markets.")
dy["A2"].font = font(italic=True, color="595959", size=9)
dy.merge_cells("A2:X2")

d_cols = [
    ("Year", 6, None),
    ("Age", 6, None),
    ("Planning years", 12, "Remaining life expectancy at this age — last-survivor (joint) years "
                           "if you set Household to Couple. The dynamic withdrawal divides the "
                           "balance by this number."),
    ("Annual return", 11, "Linked from the 4% Rule tab so both strategies see the same markets."),
    ("Inflation", 10, "Also linked from the 4% Rule tab."),
    ("Price index", 10, "Cumulative inflation since year 1."),
    ("Ordinary tax rate", 11, "Your effective rate for this phase of retirement."),
    ("Taxable beginning", 14, "Brokerage balance. Spent first."),
    ("Traditional beginning", 14, "Tax-deferred balance. Spent second and the base for RMDs."),
    ("Roth beginning", 13, "Tax-free balance. Spent last."),
    ("Taxable cost basis", 13, "Value above this is taxable gain when sold."),
    ("Total beginning balance", 15, "The three accounts combined. This is what the strategy "
                                    "divides by your planning years."),
    ("Life-exp. withdrawal", 14, "The raw actuarial amount: total balance divided by planning "
                                 "years, before guardrails or smoothing."),
    ("After guardrails", 13, "Capped inside your floor/ceiling band. The ceiling that applies at "
                             "each age is shown in the last column."),
    ("After shock absorber", 14, "Limited to your maximum year-to-year change. NOTE: this step "
                                 "can push the withdrawal back outside the guardrail band — the "
                                 "shock absorber wins, which is what the article intends."),
    ("Gross withdrawal", 14, "The amount actually taken for spending this year."),
    ("Implied rate", 10, "Gross withdrawal as a percent of the total beginning balance."),
    ("From taxable", 12, None),
    ("From traditional", 13, None),
    ("From Roth", 11, None),
    ("RMD", 11, "Required minimum distribution from the traditional balance."),
    ("RMD forced extra", 13, "The part of the RMD you did not want to spend. It is taxed and the "
                             "remainder is reinvested in the taxable account."),
    ("Tax on withdrawal", 13, "Capital-gains tax on the taxable gain plus ordinary tax on the "
                              "traditional dollars. Roth withdrawals are untaxed."),
    ("Dividend tax", 11, "Tax on the taxable account's dividends, paid every year."),
    ("Social Security", 12, "Nominal benefit, zero before your start age."),
    ("Tax on Social Security", 13, "85% of the benefit is treated as taxable."),
    ("After-tax income", 14, "Withdrawal plus Social Security, both after tax."),
    ("After-tax income (today's $)", 16, "The same figure in today's buying power."),
    ("Spending need (today's $)", 16, "Essentials, pre-65 healthcare, and long-term care in the "
                                      "years it applies."),
    ("Surplus / (shortfall)", 15, "After-tax income minus spending need, in today's dollars."),
    ("Taxable ending", 14, None),
    ("Traditional ending", 14, None),
    ("Roth ending", 12, None),
    ("Basis ending", 12, None),
    ("Total ending balance", 15, "Carried into next year."),
    ("Ending balance (today's $)", 16, "Ending balance in today's buying power."),
    ("Guardrail ceiling used", 14, "The ceiling that applied at this age. With the "
                                   "age-graduated basis this rises as you get older, so the "
                                   "life-expectancy rule keeps working instead of being "
                                   "progressively overridden by a flat cap."),
]
for i, (name, width, note) in enumerate(d_cols, start=1):
    hc = dy.cell(4, i)
    header_cell(hc, name)
    if note:
        cm = Comment(note, "Planner")
        cm.width, cm.height = 285, 120
        hc.comment = cm
    dy.column_dimensions[get_column_letter(i)].width = width
dy.row_dimensions[4].height = 42

for k in range(spec.TABLE_ROWS):
    r_ = R0 + k
    yr = k + 1
    p = r_ - 1
    dy.cell(r_, 1, yr).font = font()
    dy.cell(r_, 2, f"={q(start_age)}+{yr}-1").font = font()
    c = dy.cell(r_, 3, "=" + planning_years(f"B{r_}"))
    c.font = font()
    c.number_format = NUM1
    c = dy.cell(r_, 4, f"='4% Rule'!C{r_}")
    c.font = font(color=GREEN)
    c.number_format = PCT
    c = dy.cell(r_, 5, f"='4% Rule'!D{r_}")
    c.font = font(color=GREEN)
    c.number_format = PCT
    c = dy.cell(r_, 6, f"='4% Rule'!E{r_}")
    c.font = font(color=GREEN)
    c.number_format = "0.000"
    c = dy.cell(r_, 7, "=" + ord_rate(f"B{r_}"))
    c.font = font(color="808080")
    c.number_format = PCT
    dy.cell(r_, 8, f"={q(taxable0)}" if k == 0 else f"=MAX(AE{p},0)").font = font()
    dy.cell(r_, 9, f"={q(traditional0)}" if k == 0 else f"=MAX(AF{p},0)").font = font()
    dy.cell(r_, 10, f"={q(roth0)}" if k == 0 else f"=MAX(AG{p},0)").font = font()
    dy.cell(r_, 11, f"={q(taxable0)}*{q(basis_share)}" if k == 0 else f"=MAX(AH{p},0)"
            ).font = font(color="808080")
    dy.cell(r_, 12, f"=H{r_}+I{r_}+J{r_}").font = font()
    dy.cell(r_, 13, f"=L{r_}/MAX(C{r_},0.5)").font = font()
    dy.cell(r_, 14,
            f"=MIN({ceiling_at(f'B{r_}')}*L{r_},MAX({q(floor)}*L{r_},M{r_}))").font = font()
    if k == 0:
        dy.cell(r_, 15, f"=N{r_}")
    else:
        # "Real" applies the +/- band after restating last year's withdrawal in this
        # year's dollars, so a 5% cut is a 5% cut in buying power rather than 7.8%.
        step = f'IF({q(shock_basis)}="Real",(1+E{p}),1)'
        prev = f"P{p}*{step}"
        dy.cell(r_, 15, f"=MIN({prev}*(1+{q(shock)}),MAX({prev}*(1-{q(shock)}),N{r_}))")
    dy.cell(r_, 16, f"=MIN(O{r_},L{r_})").font = font(bold=True)
    c = dy.cell(r_, 17, f"=IFERROR(P{r_}/L{r_},0)")
    c.font = font()
    c.number_format = PCT2
    dy.cell(r_, 18, "=" + from_taxable(f"P{r_}", f"H{r_}")).font = font()
    dy.cell(r_, 19, "=" + from_traditional(f"P{r_}", f"H{r_}", f"I{r_}")).font = font()
    dy.cell(r_, 20, "=" + from_roth(f"P{r_}", f"H{r_}", f"I{r_}")).font = font()
    dy.cell(r_, 21, "=" + rmd_from(f"B{r_}", f"I{r_}")).font = font(color="808080")
    dy.cell(r_, 22, f"=MAX(MIN(U{r_},I{r_})-S{r_},0)").font = font(color="808080")
    gain = gain_fraction(f"H{r_}", f"K{r_}")
    dy.cell(r_, 23, f"=R{r_}*{gain}*{q(cg_rate)}+(S{r_}+V{r_})*G{r_}").font = font()
    dy.cell(r_, 24, f"=MAX(H{r_}-R{r_},0)*{q(div_yield)}*{q(cg_rate)}").font = font(color="808080")
    dy.cell(r_, 25, "=" + ss_gross(f"B{r_}", f"F{r_}")).font = font(color="808080")
    dy.cell(r_, 26, f"=Y{r_}*{spec.SS_TAXABLE_SHARE}*G{r_}").font = font()
    dy.cell(r_, 27,
            f"=P{r_}-(R{r_}*{gain}*{q(cg_rate)}+S{r_}*G{r_})+Y{r_}-Z{r_}").font = font(bold=True)
    dy.cell(r_, 28, f"=AA{r_}/F{r_}").font = font(color="595959")
    dy.cell(r_, 29, "=" + need_today(f"A{r_}", f"B{r_}")).font = font(color="595959")
    dy.cell(r_, 30, f"=AB{r_}-AC{r_}").font = font()
    g = f"((1+D{r_})*(1-{q(fee)}))"
    reinvest = f"V{r_}*(1-G{r_})"
    div = f"MAX(H{r_}-R{r_},0)*{q(div_yield)}"
    dy.cell(r_, 31, f"=MAX((H{r_}-R{r_}+{reinvest})*{g}-X{r_},0)").font = font()
    dy.cell(r_, 32, f"=MAX((I{r_}-S{r_}-V{r_})*{g},0)").font = font()
    dy.cell(r_, 33, f"=MAX((J{r_}-T{r_})*{g},0)").font = font()
    dy.cell(r_, 34,
            f"=MAX(K{r_}*IF(H{r_}>0,1-R{r_}/H{r_},0)+{reinvest}+{div},0)").font = font(
                color="808080")
    dy.cell(r_, 35, f"=AE{r_}+AF{r_}+AG{r_}").font = font()
    dy.cell(r_, 36, f"=AI{r_}/(F{r_}*(1+E{r_}))").font = font(color="595959")
    cy = dy.cell(r_, 37, "=" + ceiling_at(f"B{r_}"))
    cy.font = font(color=GREEN)
    cy.number_format = PCT2
    for col in range(1, 38):
        cell = dy.cell(r_, col)
        if col in (8, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                   27, 28, 29, 30, 31, 32, 33, 34, 35, 36):
            cell.number_format = CUR
        if k % 2 == 1:
            cell.fill = PatternFill("solid", start_color=LIGHT_GREY)

dy.conditional_formatting.add(f"AI{R0}:AI{RN}", red_rule)
dy.conditional_formatting.add(
    f"AD{R0}:AD{RN}",
    FormulaRule(formula=[f"AND(AD{R0}<0,A{R0}<={HORIZON})"],
                fill=PatternFill("solid", start_color=RED_FILL),
                font=Font(name=ARIAL, color=RED_FONT)))

# ===================================================================
# SHEET: Chart Data (hidden) — trims the series to the plan horizon
# ===================================================================
cd = wb.create_sheet("Chart Data")
cd.sheet_view.showGridLines = False
title(cd, "A1", "Chart series (trimmed to the plan horizon)", 12)
cd["A2"] = ("Values past the plan-to age deliberately return #N/A so the chart lines stop there. "
            "That is expected and is not a formula error.")
cd["A2"].font = font(italic=True, color="808080", size=8)
for i, name in enumerate(["Age", "4% balance (today's $)", "Dynamic balance (today's $)",
                          "4% after-tax income (today's $)",
                          "Dynamic after-tax income (today's $)",
                          "Spending need (today's $)"], start=1):
    header_cell(cd.cell(4, i), name)
    cd.column_dimensions[get_column_letter(i)].width = 20
for k in range(spec.TABLE_ROWS):
    r_ = R0 + k
    cd.cell(r_, 1, f"={q(start_age)}+{k}").font = font()
    guard = f"IF({k + 1}>{HORIZON},NA(),"
    cd.cell(r_, 2, f"={guard}'4% Rule'!AF{r_})").number_format = CUR
    cd.cell(r_, 3, f"={guard}'Dynamic Strategy'!AJ{r_})").number_format = CUR
    cd.cell(r_, 4, f"={guard}'4% Rule'!X{r_})").number_format = CUR
    cd.cell(r_, 5, f"={guard}'Dynamic Strategy'!AB{r_})").number_format = CUR
    cd.cell(r_, 6, f"={guard}'Dynamic Strategy'!AC{r_})").number_format = CUR
cd.sheet_state = "hidden"

# ===================================================================
# SHEET: Historical Returns (returns AND inflation from the same year)
# ===================================================================
hist = wb.create_sheet("Historical Returns")
hist.sheet_view.showGridLines = False
title(hist, "A1", "Historical Annual Returns & Inflation — Bootstrap Source", 14)
hist["A2"] = ("The historical Monte Carlo method samples a complete year from this table: stock "
              "return, bond return AND that year's inflation together. Keeping them paired is "
              "what makes the 1970s bite — high inflation arrived with weak real returns.")
hist["A2"].font = font(italic=True, color="595959", size=9)
hist.merge_cells("A2:E2")
hist["A3"] = ("Sources: Aswath Damodaran (NYU Stern), Annual Returns on Investments 1928-2025; "
              "US Bureau of Labor Statistics, CPI-U.")
hist["A3"].font = font(italic=True, color="808080", size=8)
hist["A3"].hyperlink = "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/histretSP.html"
hist.merge_cells("A3:E3")
for col, width in (("A", 9), ("B", 16), ("C", 20), ("D", 14), ("E", 20)):
    hist.column_dimensions[col].width = width
for col, name in enumerate(("Year", "S&P 500 return", "10-year Treasury return", "Inflation",
                            "Blended portfolio return"), start=1):
    header_cell(hist.cell(5, col), name)
HIST_FIRST = 6
mc_stock_ref = "'Monte Carlo'!$C$10"
mc_bond_ref = "'Monte Carlo'!$C$11"
for off, (year_, s_, b_, cpi_) in enumerate(spec.HISTORICAL, start=HIST_FIRST):
    hist.cell(off, 1, year_).number_format = NUM
    hist.cell(off, 1).font = font(color=BLUE)
    for col, val in ((2, s_), (3, b_), (4, cpi_)):
        c = hist.cell(off, col, val)
        c.font = font(color=BLUE)
        c.number_format = PCT
    c = hist.cell(off, 5, f"={mc_stock_ref}*B{off}+{mc_bond_ref}*C{off}")
    c.font = font(color=GREEN)
    c.number_format = PCT
HIST_LAST = HIST_FIRST + len(spec.HISTORICAL) - 1
HIST_RET = f"'Historical Returns'!$E${HIST_FIRST}:$E${HIST_LAST}"
HIST_CPI = f"'Historical Returns'!$D${HIST_FIRST}:$D${HIST_LAST}"

# ===================================================================
# SHEET: Monte Carlo (assumptions + results)
# ===================================================================
mc = wb.create_sheet("Monte Carlo")
mc.sheet_view.showGridLines = False
for col, width in (("A", 3), ("B", 50), ("C", 18), ("D", 3), ("E", 52), ("F", 16)):
    mc.column_dimensions[col].width = width
title(mc, "B2", "Monte Carlo Pressure Test", 16)
mc["B3"] = ("Both strategies run through the same randomised paths, with tax, fees, RMDs, "
            "healthcare and long-term care applied exactly as on the deterministic tabs.")
mc["B3"].font = font(italic=True, color="595959", size=9)
mc.merge_cells("B3:F3")

band(mc, 5, 2, 3, "SIMULATION ASSUMPTIONS")
mc_rows = [
    ("Return method", spec.DEFAULTS["mc_method"], "General", ["Parametric", "Historical bootstrap"],
     "Parametric draws returns from your expected return and the volatility below. Historical "
     "bootstrap resamples whole years from 1928-2025 and IGNORES your expected return — check "
     "the diagnostic on the Inputs tab for the difference."),
    (f"Simulations included (100-{spec.MAX_SIMS})", spec.DEFAULTS["mc_sims"], NUM, None, None),
    ("Annual return volatility (parametric)", spec.DEFAULTS["mc_vol"], PCT, None, None),
    ("Annual inflation volatility (parametric)", spec.DEFAULTS["mc_infl_vol"], PCT, None,
     "Inflation is randomised too, not held fixed. Inflation risk — not weak nominal returns — "
     "is what ruined the 1966 retiree cohort."),
    ("US stock allocation (historical method)", spec.DEFAULTS["mc_stock"], PCT, None, None),
]
dv_sims = DataValidation(type="whole", operator="between", formula1="100",
                         formula2=str(spec.MAX_SIMS), allow_blank=False)
dv_sims.error = f"Enter a whole number from 100 to {spec.MAX_SIMS}."
dv_pct = DataValidation(type="decimal", operator="between", formula1="0", formula2="1",
                        allow_blank=False)
dv_pct.error = "Enter a percentage from 0% to 100%."
for d in (dv_sims, dv_pct):
    mc.add_data_validation(d)

for i, (label, value, fmt, choices, note) in enumerate(mc_rows):
    r_ = 6 + i
    lc = mc.cell(r_, 2, label)
    lc.font = font()
    lc.alignment = Alignment(wrap_text=True, vertical="center")
    c = mc.cell(r_, 3, value)
    c.font = font()
    c.number_format = fmt
    c.fill = PatternFill("solid", start_color=YELLOW)
    c.border = bottom_only
    c.alignment = Alignment(horizontal="right")
    c.protection = UNLOCK
    if choices:
        d = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=False)
        d.error = "Choose one of: " + ", ".join(choices)
        mc.add_data_validation(d)
        d.add(f"C{r_}")
    elif fmt == NUM:
        dv_sims.add(f"C{r_}")
    else:
        dv_pct.add(f"C{r_}")
    if note:
        cm = Comment(note, "Planner")
        cm.width, cm.height = 300, 120
        lc.comment = cm

mc.cell(11, 2, "US Treasury bond allocation").font = font()
c = mc.cell(11, 3, "=1-C10")
c.font = font()
c.number_format = PCT
c.border = bottom_only
c.alignment = Alignment(horizontal="right")

lc = mc.cell(12, 2, "Chance of the long-term care event occurring")
lc.font = font()
lc.alignment = Alignment(wrap_text=True, vertical="center")
c = mc.cell(12, 3, spec.DEFAULTS["mc_ltc_prob"])
c.font = font()
c.number_format = PCT
c.fill = PatternFill("solid", start_color=YELLOW)
c.border = bottom_only
c.alignment = Alignment(horizontal="right")
c.protection = UNLOCK
dv_pct.add("C12")
cm = Comment("On the deterministic tabs the care event either happens or it does not. Here it "
             "happens in this share of the simulated paths, which is closer to how the risk "
             "actually behaves. Cost, duration and starting age are set on the Inputs tab.",
             "Planner")
cm.width, cm.height = 300, 120
lc.comment = cm

mc.cell(13, 2, "Fixed random batch seed").font = font()
c = mc.cell(13, 3, spec.MC_SEED)
c.font = font(color="595959")
c.number_format = NUM
c.border = bottom_only
c.alignment = Alignment(horizontal="right")

MC_METHOD = "'Monte Carlo'!$C$6"
MC_SIMS = "'Monte Carlo'!$C$7"
MC_VOL = "'Monte Carlo'!$C$8"
MC_IVOL = "'Monte Carlo'!$C$9"
MC_STOCK = "'Monte Carlo'!$C$10"
MC_BOND = "'Monte Carlo'!$C$11"
MC_LTCP = "'Monte Carlo'!$C$12"

# ===================================================================
# SHEETS: MC Engine + MC Outcomes (very hidden)
# ===================================================================
engine = wb.create_sheet("MC Engine")
engine.sheet_view.showGridLines = False
eng_headers = [
    "Simulation", "Year", "Age", "Return", "Inflation", "Price index", "Planning years",
    "Ordinary tax rate", "Spending need (today's $)",
    "4% taxable", "4% traditional", "4% Roth", "4% basis", "4% total", "4% target",
    "4% withdrawal", "4% after-tax (today's $)",
    "Dyn taxable", "Dyn traditional", "Dyn Roth", "Dyn basis", "Dyn total",
    "Dyn withdrawal", "Dyn after-tax (today's $)",
    "Return draw", "Inflation draw", "Historical row", "LTC draw", "Dyn withdrawal change",
]
for col, name in enumerate(eng_headers, start=1):
    header_cell(engine.cell(4, col), name)

E0 = 5
MC_POS = f"MIN(MAX({q(plan_age)}-{q(start_age)},1),{spec.TABLE_ROWS})"

rng = random.Random(spec.MC_SEED)
draws = []
for _ in range(spec.MAX_SIMS):
    ltc_u = rng.random()
    yrs = [(rng.gauss(0, 1), rng.gauss(0, 1), rng.randrange(1, len(spec.HISTORICAL) + 1))
           for _ in range(spec.TABLE_ROWS)]
    draws.append((ltc_u, yrs))

# MC need includes the care event only in the paths where it lands
mc_need = (f"{q(essential)}*(1+{q(drift)})^(B{{r}}-1)"
           f"+IF(C{{r}}<65,{q(healthcare)},0)"
           f"+IF(AND({q(ltc_on)}=\"Yes\",AB{{r}}<{MC_LTCP},C{{r}}>={q(ltc_age)},"
           f"C{{r}}<{q(ltc_age)}+{q(ltc_years)}),{q(ltc_cost)},0)")


def roll_accounts(engine_ws, row, prev, tx, td, rt, bs, w, cols, ord_col="H"):
    """Write the four account-state cells for a year after the first.

    The previous row's balances, withdrawal and tax rate fully determine this
    year's opening balances, so the state carries forward without needing
    separate ending-balance columns in the 71,000-row engine.
    """
    gp = f"((1+D{prev})*(1-{q(fee)}))"
    f_tx = from_taxable(f"{w}{prev}", f"{tx}{prev}")
    f_td = from_traditional(f"{w}{prev}", f"{tx}{prev}", f"{td}{prev}")
    f_rt = from_roth(f"{w}{prev}", f"{tx}{prev}", f"{td}{prev}")
    rmd_extra = f"MAX(MIN({rmd_from(f'C{prev}', f'{td}{prev}')},{td}{prev})-{f_td},0)"
    reinv = f"{rmd_extra}*(1-{ord_col}{prev})"
    div = f"MAX({tx}{prev}-{f_tx},0)*{q(div_yield)}"
    engine_ws.cell(row, cols[0],
                   f"=MAX(({tx}{prev}-{f_tx}+{reinv})*{gp}-{div}*{q(cg_rate)},0)")
    engine_ws.cell(row, cols[1], f"=MAX(({td}{prev}-{f_td}-{rmd_extra})*{gp},0)")
    engine_ws.cell(row, cols[2], f"=MAX(({rt}{prev}-{f_rt})*{gp},0)")
    engine_ws.cell(row, cols[3],
                   f"=MAX({bs}{prev}*IF({tx}{prev}>0,1-{f_tx}/{tx}{prev},0)+{reinv}+{div},0)")


def after_tax_income(row, tx, td, bs, w, ord_col="H"):
    """Spendable income: withdrawal less its tax, plus Social Security less its tax.
    The RMD excess is withdrawn but not spent, so its tax is excluded here."""
    f_tx = from_taxable(f"{w}{row}", f"{tx}{row}")
    f_td = from_traditional(f"{w}{row}", f"{tx}{row}", f"{td}{row}")
    gain = gain_fraction(f"{tx}{row}", f"{bs}{row}")
    ss = ss_gross(f"C{row}", f"F{row}")
    return (f"=({w}{row}-({f_tx}*{gain}*{q(cg_rate)}+{f_td}*{ord_col}{row})"
            f"+{ss}*(1-{spec.SS_TAXABLE_SHARE}*{ord_col}{row}))/F{row}")


for sim in range(1, spec.MAX_SIMS + 1):
    ltc_u, yrs = draws[sim - 1]
    for year in range(1, spec.TABLE_ROWS + 1):
        r_ = E0 + (sim - 1) * spec.TABLE_ROWS + year - 1
        p = r_ - 1
        z_ret, z_inf, hrow = yrs[year - 1]
        engine.cell(r_, 1, sim)
        engine.cell(r_, 2, year)
        engine.cell(r_, 3, f"={q(start_age)}+B{r_}-1")
        engine.cell(r_, 25, z_ret)
        engine.cell(r_, 26, z_inf)
        engine.cell(r_, 27, hrow)
        engine.cell(r_, 28, ltc_u)
        engine.cell(r_, 4,
                    f'=IF({MC_METHOD}="Parametric",'
                    f'EXP(LN(1+{q(exp_ret)})-0.5*{MC_VOL}^2+{MC_VOL}*Y{r_})-1,'
                    f"INDEX({HIST_RET},AA{r_}))")
        engine.cell(r_, 5,
                    f'=IF({MC_METHOD}="Parametric",'
                    f'MAX({q(infl)}+{MC_IVOL}*Z{r_},{spec.INFLATION_FLOOR}),'
                    f"INDEX({HIST_CPI},AA{r_}))")
        engine.cell(r_, 6, "=1" if year == 1 else f"=F{p}*(1+E{p})")
        engine.cell(r_, 7, "=" + planning_years(f"C{r_}"))
        engine.cell(r_, 8, "=" + ord_rate(f"C{r_}"))
        engine.cell(r_, 9, "=" + mc_need.format(r=r_))

        # ---- 4% rule: J K L M state, N total, O target, P withdrawal, Q after-tax
        if year == 1:
            engine.cell(r_, 10, f"={q(taxable0)}")
            engine.cell(r_, 11, f"={q(traditional0)}")
            engine.cell(r_, 12, f"={q(roth0)}")
            engine.cell(r_, 13, f"={q(taxable0)}*{q(basis_share)}")
            engine.cell(r_, 15,
                        f"=IF({q(current_4pct)}>0,{q(current_4pct)},N{r_}*{q(rate4)})")
        else:
            roll_accounts(engine, r_, p, "J", "K", "L", "M", "P", (10, 11, 12, 13))
            engine.cell(r_, 15, f"=O{p}*(1+E{p})")
        engine.cell(r_, 14, f"=J{r_}+K{r_}+L{r_}")
        engine.cell(r_, 16, f"=MIN(O{r_},N{r_})")
        engine.cell(r_, 17, after_tax_income(r_, "J", "K", "M", "P"))

        # ---- dynamic: R S T U state, V total, W withdrawal, X after-tax
        if year == 1:
            engine.cell(r_, 18, f"={q(taxable0)}")
            engine.cell(r_, 19, f"={q(traditional0)}")
            engine.cell(r_, 20, f"={q(roth0)}")
            engine.cell(r_, 21, f"={q(taxable0)}*{q(basis_share)}")
        else:
            roll_accounts(engine, r_, p, "R", "S", "T", "U", "W", (18, 19, 20, 21))
        engine.cell(r_, 22, f"=R{r_}+S{r_}+T{r_}")
        guarded = (f"MIN({ceiling_at(f'C{r_}')}*V{r_},"
                   f"MAX({q(floor)}*V{r_},V{r_}/MAX(G{r_},0.5)))")
        if year == 1:
            engine.cell(r_, 23, f"=MIN({guarded},V{r_})")
            engine.cell(r_, 29, 0)
        else:
            step = f'IF({q(shock_basis)}="Real",(1+E{p}),1)'
            prev_w = f"W{p}*{step}"
            engine.cell(r_, 23,
                        f"=MIN(MIN({prev_w}*(1+{q(shock)}),"
                        f"MAX({prev_w}*(1-{q(shock)}),{guarded})),V{r_})")
            engine.cell(r_, 29, f"=IF(W{p}>0,W{r_}/W{p}-1,0)")
        engine.cell(r_, 24, after_tax_income(r_, "R", "S", "U", "W"))

outcomes = wb.create_sheet("MC Outcomes")
out_headers = [
    "Simulation", "4% lasts", "4% essentials met", "4% ending (today's $)",
    "4% lowest income (today's $)", "4% cumulative income (today's $)",
    "Dyn lasts", "Dyn essentials met", "Dyn ending (today's $)",
    "Dyn lowest income (today's $)", "Dyn cumulative income (today's $)",
    "Dyn worst withdrawal change", "4% share of years covered", "Dyn share of years covered",
]
for col, name in enumerate(out_headers, start=1):
    header_cell(outcomes.cell(4, col), name)
O0 = 5
for sim in range(1, spec.MAX_SIMS + 1):
    r_ = O0 + sim - 1
    f0 = E0 + (sim - 1) * spec.TABLE_ROWS
    f1 = f0 + spec.TABLE_ROWS - 1

    def rng_(col):
        return f"'MC Engine'!${col}${f0}:INDEX('MC Engine'!${col}${f0}:${col}${f1},{MC_POS})"

    # The opening balance of the year AFTER the last modelled year is that year's
    # closing balance, which is what "balance at plan-to age" means. The horizon can
    # never reach the final table row, so this lookup is always in range.
    END_POS = f"MIN({MC_POS}+1,{spec.TABLE_ROWS})"
    four_end = f"INDEX('MC Engine'!$N${f0}:$N${f1},{END_POS})"
    dyn_end = f"INDEX('MC Engine'!$V${f0}:$V${f1},{END_POS})"
    deflator = f"INDEX('MC Engine'!$F${f0}:$F${f1},{END_POS})"
    outcomes.cell(r_, 1, sim)
    outcomes.cell(r_, 2, f"=--({four_end}>0)")
    outcomes.cell(r_, 3, f"=--(SUMPRODUCT(({rng_('Q')}<{rng_('I')})*1)=0)")
    outcomes.cell(r_, 4, f"={four_end}/({deflator})")
    outcomes.cell(r_, 5, f"=MIN({rng_('Q')})")
    outcomes.cell(r_, 6, f"=SUM({rng_('Q')})")
    outcomes.cell(r_, 7, f"=--({dyn_end}>0)")
    outcomes.cell(r_, 8, f"=--(SUMPRODUCT(({rng_('X')}<{rng_('I')})*1)=0)")
    outcomes.cell(r_, 9, f"={dyn_end}/({deflator})")
    outcomes.cell(r_, 10, f"=MIN({rng_('X')})")
    outcomes.cell(r_, 11, f"=SUM({rng_('X')})")
    # year 1 has no prior withdrawal, so the change series starts at year 2
    outcomes.cell(r_, 12,
                  f"=MIN('MC Engine'!$AC${f0 + 1}:INDEX('MC Engine'!$AC${f0 + 1}:$AC${f1},"
                  f"MAX({MC_POS}-1,1)))")
    outcomes.cell(r_, 13, f"=SUMPRODUCT(({rng_('Q')}>={rng_('I')})*1)/{MC_POS}")
    outcomes.cell(r_, 14, f"=SUMPRODUCT(({rng_('X')}>={rng_('I')})*1)/{MC_POS}")
    for col in (4, 5, 6, 9, 10, 11):
        outcomes.cell(r_, col).number_format = CUR
    outcomes.cell(r_, 12).number_format = PCT
    outcomes.cell(r_, 13).number_format = PCT
    outcomes.cell(r_, 14).number_format = PCT

engine.sheet_state = "veryHidden"
outcomes.sheet_state = "veryHidden"

# ---- Monte Carlo results panel ----
band(mc, 5, 5, 6, "RESULTS")


def mc_res(row, label, formula, fmt=PCT, bold=False):
    lc = mc.cell(row, 5, label)
    lc.font = font(bold=bold)
    lc.alignment = Alignment(wrap_text=True, vertical="center")
    c = mc.cell(row, 6, formula)
    c.font = font(bold=bold)
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    c.border = bottom_only
    return c


def oc(col):
    return f"'MC Outcomes'!{col}{O0}:INDEX('MC Outcomes'!{col}{O0}:{col}{O0 + spec.MAX_SIMS - 1},{MC_SIMS})"


mc.cell(6, 5, "— 4% RULE —").font = font(bold=True, color=MS_BLUE)
mc_res(7, "Covers essential spending EVERY year", f"=AVERAGE({oc('C')})", PCT, True)
mc_res(8, "Median share of years essentials are covered", f"=MEDIAN({oc('M')})", PCT, True)
mc_res(9, "Portfolio lasts to plan-to age", f"=AVERAGE({oc('B')})")
mc_res(10, "Median lowest year's after-tax income (today's $)", f"=MEDIAN({oc('E')})", CUR)
mc_res(11, "Median lifetime after-tax income (today's $)", f"=MEDIAN({oc('F')})", CUR)
mc_res(12, "10th percentile ending balance (today's $)", f"=PERCENTILE({oc('D')},0.1)", CUR)
mc_res(13, "Median ending balance (today's $)", f"=MEDIAN({oc('D')})", CUR)

mc.cell(16, 5, "— DYNAMIC STRATEGY —").font = font(bold=True, color=MS_BLUE)
mc_res(17, "Covers essential spending EVERY year", f"=AVERAGE({oc('H')})", PCT, True)
mc_res(18, "Median share of years essentials are covered", f"=MEDIAN({oc('N')})", PCT, True)
mc_res(19, "Portfolio lasts to plan-to age", f"=AVERAGE({oc('G')})")
mc_res(20, "Median lowest year's after-tax income (today's $)", f"=MEDIAN({oc('J')})", CUR)
mc_res(21, "Median lifetime after-tax income (today's $)", f"=MEDIAN({oc('K')})", CUR)
mc_res(22, "10th percentile ending balance (today's $)", f"=PERCENTILE({oc('I')},0.1)", CUR)
mc_res(23, "Median ending balance (today's $)", f"=MEDIAN({oc('I')})", CUR)
mc_res(24, "Median worst one-year withdrawal cut", f"=MEDIAN({oc('L')})", PCT)

for cell_ref in ("F7", "F8", "F17", "F18"):
    mc.conditional_formatting.add(cell_ref, CellIsRule(
        operator="greaterThanOrEqual", formula=["0.9"],
        fill=PatternFill("solid", start_color=GREEN_FILL),
        font=Font(name=ARIAL, bold=True, color=GREEN_FONT)))
    mc.conditional_formatting.add(cell_ref, CellIsRule(
        operator="between", formula=["0.75", "0.9"],
        fill=PatternFill("solid", start_color=AMBER_FILL),
        font=Font(name=ARIAL, bold=True, color=AMBER_FONT)))
    mc.conditional_formatting.add(cell_ref, CellIsRule(
        operator="lessThan", formula=["0.75"],
        fill=PatternFill("solid", start_color=RED_FILL),
        font=Font(name=ARIAL, bold=True, color=RED_FONT)))

mc["E26"] = ("Read the FIRST TWO lines of each block. 'Portfolio lasts' is a weak test for the "
             "dynamic strategy — a rule that spends a percentage of whatever is left can almost "
             "never reach zero, so it scores well by construction. Whether your essentials are "
             "actually covered is the question that matters, and the share-of-years line tells "
             "you how near a miss it is.")
mc["E26"].font = font(italic=True, color="595959", size=9)
mc["E26"].alignment = Alignment(wrap_text=True, vertical="top")
mc.merge_cells("E26:F29")

mc["B15"] = ("The random batch is fixed, so changing one assumption is compared against the same "
             "paths. Change MC_SEED in model_spec.py for an independent batch. Results are "
             "estimates, not guarantees: a 90% result means 900 of 1,000 modelled paths worked "
             "under these assumptions.")
mc["B15"].font = font(italic=True, color="808080", size=9)
mc["B15"].alignment = Alignment(wrap_text=True, vertical="top")
mc.merge_cells("B15:C19")

# ===================================================================
# Inputs & Summary: RESULTS panel, diagnostics, verdict, charts
# ===================================================================
band(ws, PANEL_TOP, 5, 6, "RESULTS")
RES = {}
_rrow = [PANEL_TOP + 1]


def res(label, formula, fmt=CUR, color=GREEN, section=False, bold=False, key=None,
        space=False, wrap=False, note=None):
    """Write one result row. `space` inserts a blank separator row above it;
    `wrap` lets a long text answer wrap inside the value cell; `note` attaches a
    hover explanation to the label."""
    if space:
        _rrow[0] += 1
    r_ = _rrow[0]
    lc = ws.cell(r_, 5, label)
    if section:
        lc.font = font(bold=True, color=MS_BLUE)
    else:
        lc.font = font(bold=bold)
        lc.alignment = Alignment(wrap_text=True, vertical="center")
        c = ws.cell(r_, 6, formula)
        c.font = font(color=color, bold=bold)
        if fmt != "General":
            c.number_format = fmt
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap)
        c.border = bottom_only
    if note:
        cm = Comment(note, "Planner")
        cm.width, cm.height = 320, 140
        lc.comment = cm
    if key:
        RES[key] = r_
    _rrow[0] += 1
    return r_


D_ = "'Dynamic Strategy'!"
F_ = "'4% Rule'!"


def idx(sheet, col, pos=None):
    return f"INDEX({sheet}${col}${R0}:${col}${RN},{pos or HORIZON})"


def rng_to_horizon(sheet, col):
    return f"{sheet}${col}${R0}:{idx(sheet, col)}"


res("Years modelled (start age to plan-to age)", f"={q(plan_age)}-{q(start_age)}", NUM, BLACK,
    key="years")
res("Net expected return after fees",
    f"=(1+{q(exp_ret)})*(1-{q(fee)})-1", PCT2, BLACK, key="netret", space=True)

res("— DOES THE PLAN WORK? —", None, section=True)
res("Dynamic: essentials covered every year?",
    f'=IF(SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*({D_}$AD${R0}:$AD${RN}<0))=0,"YES","NO")',
    "General", BLACK, bold=True, key="dyn_ess")
res("4% rule: essentials covered every year?",
    f'=IF(SUMPRODUCT(({F_}$A${R0}:$A${RN}<={HORIZON})*({F_}$Z${R0}:$Z${RN}<0))=0,"YES","NO")',
    "General", BLACK, bold=True, key="four_ess")

res("— DYNAMIC STRATEGY —", None, section=True, space=True)
res("First-year after-tax income (today's $)", f"={D_}$AB${R0}", CUR, GREEN, key="dyn_first")
res("Lowest annual after-tax income (today's $)",
    f"=MIN({rng_to_horizon(D_, 'AB')})", CUR, GREEN, key="dyn_low")
res("Highest annual after-tax income (today's $)",
    f"=MAX({rng_to_horizon(D_, 'AB')})", CUR, GREEN, key="dyn_high")
res("Lifetime after-tax income (today's $)",
    f"=SUM({rng_to_horizon(D_, 'AB')})", CUR, GREEN, key="dyn_total")
res("Years spending falls short of needs",
    f"=SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*({D_}$AD${R0}:$AD${RN}<0))",
    NUM, BLACK, key="dyn_short")
res("Share of years essentials are covered",
    f"=1-SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*({D_}$AD${R0}:$AD${RN}<0))/{HORIZON}",
    PCT, BLACK, key="dyn_share")
res("Real income at plan-to age vs. year 1",
    f"={idx(D_, 'AB')}/{D_}$AB${R0}-1", PCT, BLACK, key="dyn_decay")
res("Balance at plan-to age (today's $)", f"={idx(D_, 'AJ')}", CUR, GREEN, key="dyn_bal")
res("Money lasts to plan-to age?", f'=IF({idx(D_, "AI")}>0,"YES","NO")', "General", BLACK,
    key="dyn_lasts")
res("Guardrail ceiling starts capping spending at",
    f'=IF(MIN({BIND_RANGE})>=9999,"never — the rule runs freely",'
    f'"age "&MIN({BIND_RANGE})&" (of "&{HORIZON}&" years modelled)")',
    "General", BLACK, key="ceil_bind_age", wrap=True)


def exhausted_at(col):
    """The age at which one account's ending balance first reaches zero."""
    rng = f"{D_}${col}${R0}:{idx(D_, col)}"
    return (f'=IF(COUNTIF({rng},0)=0,"not within the plan",'
            f'"age "&({q(start_age)}+MATCH(0,{rng},0)-1))')


# The withdrawal order is brokerage -> traditional -> Roth, so these three ages map
# the whole tax arc of the plan: cheap, then expensive, then free.
res("Brokerage accounts exhausted at", exhausted_at("AE"), "General", BLACK,
    key="tx_dry",
    note="Spent first. Only the GAIN is taxed, at capital-gains rates, so these are your "
         "cheapest withdrawals. When this account empties you move to the traditional IRA "
         "and your tax bill RISES. It is also the point at which you lose the flexibility "
         "to control which year your taxable income lands in — which is what makes Roth "
         "conversions and ACA-subsidy management possible.")
res("Traditional IRA accounts exhausted at", exhausted_at("AF"), "General", BLACK,
    key="td_dry",
    note="Spent second. Every dollar is ordinary income — the deferred tax bill coming due — "
         "and this is the balance RMDs are calculated from. When it empties you move to the "
         "Roth and your withdrawals become tax-free.")
res("Roth accounts exhausted at", exhausted_at("AG"), "General", BLACK,
    key="roth_dry",
    note="Spent last, so it compounds untouched for as long as possible. Withdrawals are "
         "tax-free and it is not subject to RMDs. If this shows an age inside your plan, the "
         "portfolio has run out entirely.")

res("— 4% RULE —", None, section=True, space=True)
res("First-year after-tax income (today's $)", f"={F_}$X${R0}", CUR, GREEN, key="four_first")
res("Lowest annual after-tax income (today's $)",
    f"=MIN({rng_to_horizon(F_, 'X')})", CUR, GREEN, key="four_low")
res("Lifetime after-tax income (today's $)",
    f"=SUM({rng_to_horizon(F_, 'X')})", CUR, GREEN, key="four_total")
res("Years spending falls short of needs",
    f"=SUMPRODUCT(({F_}$A${R0}:$A${RN}<={HORIZON})*({F_}$Z${R0}:$Z${RN}<0))",
    NUM, BLACK, key="four_short")
res("Share of years essentials are covered",
    f"=1-SUMPRODUCT(({F_}$A${R0}:$A${RN}<={HORIZON})*({F_}$Z${R0}:$Z${RN}<0))/{HORIZON}",
    PCT, BLACK, key="four_share")
res("Balance at plan-to age (today's $)", f"={idx(F_, 'AF')}", CUR, GREEN, key="four_bal")
res("Money lasts to plan-to age?", f'=IF({idx(F_, "AE")}>0,"YES","NO")', "General", BLACK,
    key="four_lasts")

res("— WHAT YOUR ESSENTIALS COST —", None, section=True, space=True)
res("Social Security you will actually receive (today's $/yr, after tax)",
    f"={q(full_ss)}*{SS_FACTOR}*(1-{spec.SS_TAXABLE_SHARE}*{q(tax_ss)})", CUR, BLACK,
    key="ss_net")
res("Essentials the portfolio must cover once collecting (today's $/yr)",
    f"=MAX({q(essential)}-{q(full_ss)}*{SS_FACTOR}*(1-{spec.SS_TAXABLE_SHARE}*{q(tax_ss)}),0)",
    CUR, BLACK, key="ess_gap")

for key in ("dyn_ess", "four_ess", "dyn_lasts", "four_lasts"):
    r_ = RES[key]
    ws.conditional_formatting.add(f"F{r_}", CellIsRule(
        operator="equal", formula=['"YES"'],
        fill=PatternFill("solid", start_color=GREEN_FILL),
        font=Font(name=ARIAL, bold=True, color=GREEN_FONT)))
    ws.conditional_formatting.add(f"F{r_}", CellIsRule(
        operator="equal", formula=['"NO"'],
        fill=PatternFill("solid", start_color=RED_FILL),
        font=Font(name=ARIAL, bold=True, color=RED_FONT)))
for key in ("dyn_short", "four_short"):
    r_ = RES[key]
    ws.conditional_formatting.add(f"F{r_}", CellIsRule(
        operator="equal", formula=["0"],
        fill=PatternFill("solid", start_color=GREEN_FILL),
        font=Font(name=ARIAL, bold=True, color=GREEN_FONT)))
    ws.conditional_formatting.add(f"F{r_}", CellIsRule(
        operator="greaterThanOrEqual", formula=["1"],
        fill=PatternFill("solid", start_color=RED_FILL),
        font=Font(name=ARIAL, bold=True, color=RED_FONT)))

# ---- verdict box ----
TOTAL0 = f"({q(taxable0)}+{q(traditional0)}+{q(roth0)})"
vcell.value = (
    f'="Starting at age "&{q(start_age)}&" with "&TEXT({TOTAL0},"$#,##0")&'
    f'", planning to age "&{q(plan_age)}&".     '
    f'DYNAMIC: essentials covered every year? "&$F${RES["dyn_ess"]}&'
    f'" — after-tax income "&TEXT($F${RES["dyn_low"]},"$#,##0")&" to "&'
    f'TEXT($F${RES["dyn_high"]},"$#,##0")&" a year in today''s dollars.     '
    f'4% RULE: essentials covered every year? "&$F${RES["four_ess"]}&'
    f'" — lowest year "&TEXT($F${RES["four_low"]},"$#,##0")&'
    f'".     All figures are after tax and after fees."'
)

# ---- diagnostics ----
DIAG_ROW = _rrow[0] + 1
band(ws, DIAG_ROW, 5, 6, "PLAN DIAGNOSTICS — read these before trusting the numbers")
CEIL_BINDS = (f"SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*"
              f"({D_}$M${R0}:$M${RN}>{D_}$AK${R0}:$AK${RN}*{D_}$L${R0}:$L${RN}+0.01))")
FIRST_BIND = f"MIN({BIND_RANGE})"
diagnostics = [
    (f'=IF({q(ceiling_basis)}="Fixed %",'
     f'IF({FIRST_BIND}>=9999,'
     f'"OK: your flat "&TEXT({q(ceiling)},"0.0%")&" ceiling never caps the life-expectancy '
     f'withdrawal over this horizon.",'
     f'"NOTE: a flat ceiling is age-blind. Yours caps the life-expectancy withdrawal from age "'
     f'&{FIRST_BIND}&" onward — "&{CEIL_BINDS}&" of "&{HORIZON}&" years — so for those years you '
     f'are no longer following the life-expectancy strategy at all; you are simply spending "'
     f'&TEXT({q(ceiling)},"0.0%")&" of the balance. Switch the ceiling basis to Age-graduated to '
     f'let the strategy keep working."),'
     f'IF({FIRST_BIND}>=9999,'
     f'"OK: the age-graduated ceiling never binds over this horizon, so the life-expectancy rule '
     f'runs freely. If you want a firmer brake at very old ages, lower the multiple.",'
     f'"OK: the age-graduated ceiling first caps spending at age "&{FIRST_BIND}&" ("&{CEIL_BINDS}&'
     f'" of "&{HORIZON}&" years). Before that the life-expectancy rule runs freely; after it, the '
     f'cap trims what would otherwise be an extreme withdrawal rate."))'),
    (f'=IF(AND({q(plan_age)}-{q(start_age)}>30,{q(rate4)}>=0.04),'
     f'"WARNING: you are modelling "&({q(plan_age)}-{q(start_age)})&" years, but the 4% rule was '
     f'only ever validated for 30. Over this horizon 4% is not a safe rate — treat the 4% column '
     f'as a reference point, not a plan.",'
     f'"OK: the horizon is within the range the 4% rule was tested for.")'),
    (f'=IF({q(plan_age)}-{q(start_age)}>{spec.TABLE_ROWS},'
     f'"WARNING: the horizon exceeds the {spec.TABLE_ROWS} rows in the tables and has been '
     f'truncated. Lower the plan-to age.","OK: the horizon fits inside the tables.")'),
    (f'=IF({MC_METHOD}="Historical bootstrap",'
     f'"NOTE: the Monte Carlo is using historical bootstrap, which samples 1928-2025 US returns '
     f'and IGNORES your "&TEXT({q(exp_ret)},"0.0%")&" expected return. US history averaged about '
     f'9% for a 60/40 mix, so results will look markedly better than your own assumption.",'
     f'"OK: the Monte Carlo is using your expected return.")'),
    (f'=IF({q(tax_early)}+{q(tax_ss)}+{q(tax_rmd)}=0,'
     f'"WARNING: all three effective tax rates are 0%. Unless everything is in a Roth, this '
     f'overstates spendable income substantially.","OK: taxes are being modelled.")'),
    (f'=IF({q(ltc_on)}="No",'
     f'"NOTE: no long-term care event is included. Around half of 65-year-olds will need some '
     f'paid care, and it is the largest single tail risk in most retirement plans.",'
     f'"OK: a long-term care event is included.")'),
    (f'=IF({q(household)}="Couple",'
     f'"OK: using last-survivor life expectancy, which is the right divisor for a couple.",'
     f'"NOTE: using single life expectancy. If you have a spouse or partner, switch Household to '
     f'Couple — planning to one life is the more common and more dangerous error.")'),
    (f'=IF(SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*'
     f'({D_}$Q${R0}:$Q${RN}>{D_}$AK${R0}:$AK${RN}+0.0001))>0,'
     f'"NOTE: in "&SUMPRODUCT(({D_}$A${R0}:$A${RN}<={HORIZON})*'
     f'({D_}$Q${R0}:$Q${RN}>{D_}$AK${R0}:$AK${RN}+0.0001))&" year(s) the shock absorber holds '
     f'spending ABOVE the guardrail ceiling. That is intended — the smoothing rule outranks the '
     f'band — but it does draw the portfolio down faster.",'
     f'"OK: the withdrawal rate stayed inside the guardrail band.")'),
    (f'=IF({q(taxable0)}+{q(roth0)}=0,'
     f'"WARNING: every dollar is in tax-deferred accounts, so every withdrawal is fully taxable '
     f'and RMDs apply to the whole portfolio. If you actually hold a brokerage or Roth account, '
     f'split the balances out — it materially changes after-tax income.",'
     f'IF({q(taxable0)}/({q(taxable0)}+{q(traditional0)}+{q(roth0)})<0.05,'
     f'"NOTE: you have almost nothing in a taxable account, so there is little flexibility to '
     f'manage which year income lands in.",'
     f'"OK: your balances are split across account types."))'),
    (f'=IF(AND({q(ss_age)}-{q(start_age)}>=5,{q(tax_early)}>={q(tax_ss)}),'
     f'"NOTE: you retire "&({q(ss_age)}-{q(start_age)})&" years before Social Security starts, '
     f'but your early-retirement tax rate is not lower than your later rate. Those low-income '
     f'years are usually your cheapest tax years and a prime window for Roth conversions — check '
     f'the early rate is realistic.",'
     f'"OK: the early-retirement tax rate looks consistent with the later phases.")'),
    (f'=IF($F${RES["dyn_short"]}+$F${RES["four_short"]}=0,'
     f'"OK: both strategies fund your essentials in every modelled year.",'
     f'"WARNING: at these inputs the plan does not fully fund your needs — dynamic falls short in '
     f'"&$F${RES["dyn_short"]}&" year(s), the 4% rule in "&$F${RES["four_short"]}&". The usual '
     f'causes are pre-65 healthcare, the care event, and tax. Try a later start age, lower '
     f'essentials, or a larger portfolio.")'),
]
for i, formula in enumerate(diagnostics):
    r_ = DIAG_ROW + 1 + i
    c = ws.cell(r_, 5, formula)
    c.font = font(size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    # every row that carries a diagnostic gets the taller height, set as it is
    # written so the two can never drift apart
    ws.row_dimensions[r_].height = 45
    ws.merge_cells(start_row=r_, start_column=5, end_row=r_, end_column=6)
    ws.conditional_formatting.add(f"E{r_}", FormulaRule(
        formula=[f'LEFT(E{r_},7)="WARNING"'],
        fill=PatternFill("solid", start_color=RED_FILL),
        font=Font(name=ARIAL, size=9, color=RED_FONT)))
    ws.conditional_formatting.add(f"E{r_}", FormulaRule(
        formula=[f'LEFT(E{r_},4)="NOTE"'],
        fill=PatternFill("solid", start_color=AMBER_FILL),
        font=Font(name=ARIAL, size=9, color=AMBER_FONT)))
    ws.conditional_formatting.add(f"E{r_}", FormulaRule(
        formula=[f'LEFT(E{r_},2)="OK"'],
        fill=PatternFill("solid", start_color=GREEN_FILL),
        font=Font(name=ARIAL, size=9, color=GREEN_FONT)))
DIAG_LAST = DIAG_ROW + len(diagnostics)

# ---- charts ----
CHART_ROW = max(IB.row, DIAG_LAST) + 2
ws.cell(CHART_ROW, 2, "Visual comparison").font = font(bold=True, color="242424", size=12)

c1 = LineChart()
c1.title = "Portfolio balance in today's dollars"
c1.height, c1.width = 8, 17
c1.y_axis.title = "Balance (today's $)"
c1.x_axis.title = "Age"
c1.x_axis.delete = False
c1.y_axis.delete = False
c1.series.append(Series(Reference(cd, min_col=2, min_row=R0, max_row=RN), title="4% rule"))
c1.series.append(Series(Reference(cd, min_col=3, min_row=R0, max_row=RN), title="Dynamic strategy"))
c1.set_categories(Reference(cd, min_col=1, min_row=R0, max_row=RN))
ws.add_chart(c1, f"B{CHART_ROW + 1}")

c2 = LineChart()
c2.title = "After-tax income vs. what you need (today's dollars)"
c2.height, c2.width = 8, 17
c2.y_axis.title = "Today's $ per year"
c2.x_axis.title = "Age"
c2.x_axis.delete = False
c2.y_axis.delete = False
c2.series.append(Series(Reference(cd, min_col=4, min_row=R0, max_row=RN), title="4% rule income"))
c2.series.append(Series(Reference(cd, min_col=5, min_row=R0, max_row=RN),
                        title="Dynamic strategy income"))
c2.series.append(Series(Reference(cd, min_col=6, min_row=R0, max_row=RN),
                        title="Spending need"))
c2.set_categories(Reference(cd, min_col=1, min_row=R0, max_row=RN))
ws.add_chart(c2, f"B{CHART_ROW + 17}")

FOOT = CHART_ROW + 33
ws.cell(FOOT, 2,
        "Educational illustration, not financial advice. Inspired by the WSJ article \"The 4% Rule "
        "for Retirement Is Too Simple. Here's a Better Way.\" Life-expectancy figures are SSA "
        "cohort values; RMD divisors are the IRS Uniform Lifetime Table; Social Security factors "
        "assume birth in 1960 or later.").font = font(size=8, italic=True, color="808080")
ws.merge_cells(start_row=FOOT, start_column=2, end_row=FOOT, end_column=6)

# ===================================================================
# SHEET: Instructions
# ===================================================================
ins = wb.create_sheet("Instructions")
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 112

instructions = [
    ("h1", "How to Use This Planner"),
    ("p", "This workbook compares two ways to spend down a retirement portfolio: the classic 4% "
          "rule and a flexible \"dynamic\" strategy (see the Article tab). Unlike a simple "
          "illustration, it models tax, fees, required minimum distributions, pre-Medicare "
          "healthcare and long-term care, because those are what usually decide whether a plan "
          "actually works. It is still an educational tool, not financial advice."),
    ("h2", "Start here"),
    ("p", "1. Open Inputs & Summary and work down the yellow cells. Every input has a hover note."),
    ("p", "2. Read the blue verdict box, then the RESULTS panel, then PLAN DIAGNOSTICS. The "
          "diagnostics tell you when an assumption has quietly broken the model."),
    ("p", "3. Scroll down for the two charts. The second one — after-tax income against what you "
          "need — is the one that answers 'can I afford this?'"),
    ("p", "4. Open the Monte Carlo tab and pressure-test both strategies."),
    ("h2", "The one number to look at"),
    ("p", "\"Covers essential spending every year\" is the headline metric, not \"portfolio "
          "lasts\". A strategy that withdraws a percentage of whatever is left can almost never "
          "reach zero, so it scores nearly 100% on survival no matter how badly it is doing. That "
          "flatters it. The honest question is whether the income it produces actually covers "
          "your needs in every year, which is what the essentials test measures."),
    ("h2", "Everything is after tax"),
    ("p", "Enter your essential spending as an AFTER-TAX number, because that is what you really "
          "have to fund. The model applies your effective tax rate to the tax-deferred share of "
          "each withdrawal and to 85% of your Social Security, then compares what is left with "
          "what you need. A pre-tax projection can easily overstate spendable income by 15-25%, "
          "which is enough to flip a plan from working to failing."),
    ("h2", "Required minimum distributions"),
    ("p", "From your RMD start age (73, or 75 if born in 1960 or later) the IRS forces a minimum "
          "withdrawal from tax-deferred accounts. The model treats this correctly as a tax event "
          "rather than forced spending: if the RMD exceeds what you wanted to spend, the excess "
          "comes out, gets taxed, and the remainder stays invested. Only the tax actually leaves "
          "your plan."),
    ("h2", "Life expectancy, and why couples must change one setting"),
    ("p", "The dynamic strategy divides your balance by your remaining years. Enter four anchors "
          "from the SSA calculator and the Life Expectancy tab interpolates every age, then "
          "derives a full survival curve from them."),
    ("p", "If you have a spouse or partner, set Household to Couple. The model then uses "
          "last-survivor life expectancy — the expected time until the SECOND death — which at 65 "
          "is roughly 26 years against about 20 for a single life. Planning a couple's money to "
          "one life expectancy is a common and expensive mistake."),
    ("p", "Remember that life expectancy is an average: about half of people outlive it. The "
          "dynamic rule handles this gracefully because it recalculates every year and the "
          "divisor never reaches zero, but it does mean spending is front-loaded."),
    ("h2", "Guardrails and the shock absorber"),
    ("p", "Each year the raw actuarial withdrawal is capped inside your floor/ceiling band, then "
          "limited by the shock absorber. Note the order: the shock absorber runs last, so in a "
          "sharp drawdown it can hold your spending ABOVE the ceiling. That is deliberate — it is "
          "what stops spending collapsing — but it draws the portfolio down faster, and the "
          "diagnostics tell you when it is happening."),
    ("p", "Set the shock absorber to Real (the default) so a 5% cut means 5% of buying power. On "
          "a Nominal basis with 3% inflation, a +/-5% band is really +1.9% / -7.8% in real terms, "
          "which quietly ratchets your standard of living down."),
    ("h2", "Why the ceiling is age-graduated"),
    ("p", "The actuarial withdrawal is balance divided by remaining years, so its implied RATE is "
          "always 1 / remaining years — the balance cancels out. That means the guardrail band "
          "does not react to markets at all; it is purely an age switch. The ceiling starts "
          "binding at whatever age your remaining years fall below 1 / ceiling, and from then on "
          "you are no longer following the life-expectancy strategy — you are simply spending a "
          "flat percentage. (The shock absorber is the part that responds to markets.)"),
    ("p", "With a flat 6% ceiling that switch flips at about age 71, which on a 40-year plan "
          "means the strategy you are supposedly testing is inactive for roughly half of it. So "
          "the default here instead sets the cap to a multiple of the IRS Uniform Lifetime "
          "withdrawal rate — the same published table used for RMDs, on the Reference Tables tab "
          "— which rises with age. At the default 2.0x the cap first bites around age 83, so the "
          "actuarial rule runs freely through your seventies and the ceiling only trims a "
          "genuinely extreme rate late in life."),
    ("p", "Two caveats. The IRS divisors assume a beneficiary ten years younger, so they are "
          "deliberately conservative as a withdrawal rate; the multiple is what turns them into "
          "a sensible cap. And a looser ceiling means more spending and a smaller estate — on the "
          "default assumptions, moving from a flat 6% to 2.0x adds roughly $450,000 of lifetime "
          "after-tax spending and removes roughly $640,000 from what is left at the end. Neither "
          "is right or wrong; choose the one that matches your intent. Set the basis to Fixed % "
          "if you want the original article's behaviour."),
    ("p", "The RESULTS panel reports the exact age at which your ceiling starts capping spending, "
          "and the Dynamic Strategy tab shows the ceiling used in every single year."),
    ("h2", "Market scenarios"),
    ("p", "Steady uses your expected return and inflation every year. Early crash front-loads a "
          "sharp drop to test sequence-of-returns risk. Stagflation combines weak returns with "
          "HIGH inflation — that combination, not weak nominal returns alone, is what ruined the "
          "1966 retiree cohort. You can edit every number in the Scenario Detail table."),
    ("h2", "The Monte Carlo test"),
    ("p", "Parametric mode draws returns from your expected return and volatility, and randomises "
          "inflation too. Historical bootstrap resamples complete years from 1928-2025, taking the "
          "stock return, bond return and that year's inflation together so their relationship "
          "survives."),
    ("p", "One caution the diagnostics repeat: historical bootstrap ignores your expected return. "
          "US history averaged roughly 9% a year for a 60/40 mix, far above most forward-looking "
          "assumptions, and it is the single best-performing major market of the last century. "
          "Treat those results as an optimistic bound."),
    ("h2", "Already retired? Use this every year"),
    ("p", "The dynamic strategy is designed to be re-run annually. Set the retirement start age to "
          "your current age and the portfolio to your current balance; year 1 becomes the coming "
          "year and the first-year figure is your suggested amount. Come back each year and "
          "update both. For the 4% column, enter your existing inflation-adjusted withdrawal in "
          "\"Current 4% rule withdrawal\" if you want it continued rather than restarted."),
    ("h2", "What this model still does not do"),
    ("p", "It uses one blended portfolio rather than separate traditional, Roth and taxable "
          "accounts, so it cannot model withdrawal sequencing or Roth conversions. Tax is a "
          "single effective rate, not brackets, and capital gains in taxable accounts are treated "
          "as tax-free. It ignores state-specific rules, IRMAA surcharges, survivor benefit "
          "changes, and any reduction in Social Security if the trust fund is not topped up. "
          "Every one of these omissions makes the plan look slightly BETTER than reality, so "
          "treat a marginal result as a failing one."),
    ("h2", "Colour key"),
    ("p", "Yellow = inputs you can edit.  Black = calculated.  Green = pulled from another tab.  "
          "Grey = restated in today's dollars.  Red = a shortfall or a depleted balance."),
    ("h2", "Why the sheets are locked"),
    ("p", "Every tab is protected so only the yellow cells can be changed. There is no password: "
          "Review ribbon -> Unprotect Sheet unlocks immediately."),
]
ir = 2
for kind, text in instructions:
    if kind == "h2":
        ir += 1
    c = ins.cell(ir, 2, text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if kind == "h1":
        c.font = font(bold=True, color="242424", size=18)
        ins.row_dimensions[ir].height = 26
    elif kind == "h2":
        c.font = font(bold=True, color=MS_BLUE, size=13)
        ins.row_dimensions[ir].height = 22
    else:
        c.font = font(color="242424", size=11)
        ins.row_dimensions[ir].height = 15 * max(1, (len(text) // 108) + 1) + 4
    ir += 1

# ===================================================================
# SHEET: Article
# ===================================================================
art = wb.create_sheet("Article")
art.sheet_view.showGridLines = False
art.column_dimensions["A"].width = 3
art.column_dimensions["B"].width = 112

article = [
    ("h1", "The 4% Rule for Retirement Is Too Simple. Here's a Better Way."),
    ("src", "The Wall Street Journal · Personal Finance / Retirement"),
    ("url", "https://www.wsj.com/personal-finance/retirement/retirement-planning-4-percent-rule-d5ba3a0b"),
    ("p", "Most retirees are familiar with the 4% rule."),
    ("p", "It's the guide many financial pros have extolled for decades for the decumulation "
          "phase of retirement - the time when you finally start tapping the money you've spent "
          "decades (ideally) saving and investing."),
    ("p", "Withdraw 4% of your portfolio in the first year of retirement, then adjust that amount "
          "for inflation every year thereafter. If your portfolio starts out at $1 million, for "
          "instance, you spend $40,000 the first year, then $40,000 plus inflation the next year, "
          "and so on."),
    ("p", "The rule was devised in 1994 by financial adviser William Bengen after research showed "
          "that retirees with a balanced stock-and-bond portfolio who followed that math wouldn't "
          "have run out of money over any 30-year period since 1926 - even when economic "
          "conditions were bad."),
    ("p", "It's pretty simple. Maybe too simple."),
    ("p", "The 4% rule has some blind spots. It assumes a 30-year retirement, but some retirees "
          "will need their money to last longer than that. It also assumes the markets will "
          "perform as well as they have over the past century. As a result, researchers have been "
          "lowering or raising the figure for years. Investment research firm Morningstar revises "
          "its estimate annually; it was 3.3% in 2021 and 3.9% this year. Bengen himself has said "
          "that a safe number could now be 4.7%."),
    ("p", "The rule's deeper problem is its rigidity. Your portfolio amount could double in a "
          "market boom during retirement and the rule would still limit you to the same "
          "inflation-adjusted amount, which would be unnecessarily frugal. More concerning, your "
          "assets could fall by a third and the rule wouldn't tell you to curtail your spending. "
          "You would have to liquidate a bigger portion of a shrinking portfolio, leaving you at "
          "a real risk of eventually outliving your money, despite what the historical data say."),
    ("p", "The good news is that there are ways to make retirement spending more responsive to how "
          "long you're likely to live and how markets perform. Use the 4% rule as a reference "
          "point, rather than have it be the full plan."),
    ("h2", "Your life expectancy matters"),
    ("p", "Instead of spending a fixed inflation-adjusted dollar amount every year, you need to "
          "factor in your life expectancy and the size of your portfolio and then adjust "
          "accordingly."),
    ("p", "You divide your current portfolio balance by the number of years you can expect to "
          "live, and spend that much in the coming year. The following year, you do the math "
          "again with your new portfolio balance and your new remaining life expectancy. (You can "
          "find your life expectancy at your current age in the Social Security Administration's "
          "life tables, which are updated annually. Keep in mind that tables don't apply to any "
          "specific person; they don't factor in health issues or family medical history. You may "
          "want to be conservative and go beyond your life expectancy, or take a chance and "
          "reduce it.)"),
    ("p", "If you're a 70-year-old woman, for instance, your remaining life expectancy is about 16 "
          "years. So you would divide your portfolio into sixteenths and spend one of them this "
          "year. A year later, you would recalculate using your new balance and your new life "
          "expectancy of 15 years, and so on. This allows your spending to rise when your "
          "portfolio is up and fall when it's down."),
    ("p", "This actuarial approach lessens the chance you'll run out of money, and it lets good "
          "market returns flow through to higher spending."),
    ("h2", "Spending too much - or not enough"),
    ("p", "But it introduces a new problem: too much fluctuation in your spending."),
    ("p", "If markets drop 30% in a bad year, your spending drops by roughly 30%, too. If markets "
          "surge in a year when the life tables say your remaining expectancy is five years, the "
          "formula tells you to spend a fifth of your portfolio - probably far more than you need "
          "or want."),
    ("p", "The fix is to put guardrails on the actuarial approach, using the 4% rule as your "
          "anchor. You set an upper and lower band around it - say, 3% of assets on the low end "
          "and 6% on the high end. Each year, you do the actuarial calculation and check the "
          "result against the band."),
    ("p", "If the implied withdrawal rate falls inside the guardrails, you're fine. If your "
          "spending would exceed 6% of your portfolio, you cap your spending at 6%. If your "
          "spending would fall below 3%, you spend at least 3%. The guardrails work like a "
          "financial thermostat. The 4% rule sits near the center of the band as the familiar "
          "reference point most retirees already know."),
    ("p", "But if markets fluctuate a lot, you could still be in for a shock. Say markets fall "
          "sharply and the guardrails tell you to cut spending by a quarter to get back inside "
          "the band. That could be wrenching for a lot of retirees."),
    ("p", "So here's a third fix: a shock absorber, like the ones some actual roadside guardrails "
          "have."),
    ("p", "You cap year-to-year spending changes at, say, plus or minus 5%. Even if the guardrails "
          "say to cut more aggressively, you never cut by more than 5% in a single year. Same on "
          "the upside. Your spending won't swing chaotically between feast and famine."),
    ("h2", "Essential vs. discretionary"),
    ("p", "The trade-off for the shock absorption is that in a deep, sustained downturn, your "
          "portfolio will be exposed to more drawdown than a stricter rule would allow. So the "
          "shock absorber works best for retirees with some cushion above their essential "
          "spending."),
    ("p", "That points to the next refinement."),
    ("p", "None of the changes so far distinguish between the spending you have to do and the "
          "spending you would like to do. If you feel any of these strategies leaves you too "
          "vulnerable to market swings or a surprisingly long life, you can calculate the "
          "expenses you absolutely need to cover - food, healthcare, rent and the like - and pay "
          "for those with guaranteed income such as Social Security benefits or an annuity."),
    ("p", "Then you can apply a dynamic withdrawal strategy, with the guardrails and shock "
          "absorber discussed above, to the spending that remains. You can afford to take more "
          "risk, and probably earn higher returns, with your discretionary portfolio because you "
          "know the essentials are covered."),
    ("p", "None of these refinements is perfect. Each retiree has personal factors that could "
          "change the math, and the account order above will have exceptions in particular tax "
          "situations. There are software tools that run the calculations back and forth until "
          "they converge on an answer for a given person's circumstances. Realistically, though, "
          "most people won't use such tools routinely, if at all."),
    ("p", "A dynamic strategy - with the 4% rule as a reference point - is practicable without "
          "specialized software. And it should help you enjoy your nest egg and not outlive it."),
]
ar = 2
for kind, text in article:
    if kind == "h2":
        ar += 1
    c = art.cell(ar, 2, text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if kind == "h1":
        c.font = font(bold=True, color="242424", size=18)
        art.row_dimensions[ar].height = 26
    elif kind == "src":
        c.font = font(italic=True, color="808080", size=9)
    elif kind == "url":
        c.font = Font(name=ARIAL, color="0563C1", size=10, underline="single")
        c.hyperlink = text
        ar += 1
    elif kind == "h2":
        c.font = font(bold=True, color=MS_BLUE, size=13)
        art.row_dimensions[ar].height = 22
    else:
        c.font = font(color="242424", size=11)
        art.row_dimensions[ar].height = 15 * max(1, (len(text) // 108) + 1) + 4
    ar += 1

# ===================================================================
# Tab order, protection, save
# ===================================================================
order = ["Instructions", "Inputs & Summary", "Monte Carlo", "Dynamic Strategy", "4% Rule",
         "Life Expectancy", "Historical Returns", "Reference Tables", "Article",
         "Chart Data", "MC Engine", "MC Outcomes"]
wb._sheets.sort(key=lambda s: order.index(s.title))

for sh in wb.worksheets:
    sh.protection.sheet = True
    sh.protection.selectLockedCells = False
    sh.protection.selectUnlockedCells = False
    sh.protection.formatCells = False

wb.properties.creator = "Retirement Planner"
wb.properties.lastModifiedBy = "Retirement Planner"
wb.properties.title = "Retirement Withdrawal Strategies Planner"
wb.properties.description = ("Retirement withdrawal planner: 4% rule vs. a dynamic "
                             "life-expectancy strategy, after tax, fees, RMDs and long-term care, "
                             "with Monte Carlo pressure testing.")
wb.properties.keywords = None
wb.properties.company = None

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Retirement Withdrawal Strategies Planner.xlsx")


def main():
    wb.save(OUT)
    print("saved:", OUT)


if __name__ == "__main__":
    main()
