"""Independently recompute the planner model and compare with what Excel calculated.

This is the real correctness gate. `recalc_excel.ps1` only proves no cell contains
an error value; this proves the numbers are the numbers they are supposed to be.
Every formula on the strategy tabs, the life-expectancy engine, the results panel
and all 1,000 Monte Carlo outcomes are recomputed here from `model_spec.py` and
checked against the workbook cell by cell.

    python validate_model.py
"""
import math
import os
import random
import sys

from openpyxl import load_workbook

import model_spec as spec

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, "Retirement Withdrawal Strategies Planner.xlsx")

D = dict(spec.DEFAULTS)
R0 = 5
N = spec.TABLE_ROWS
AGES = list(range(spec.LE_FIRST_AGE, spec.LE_LAST_AGE + 1))
SS_MAP = dict(spec.SS_FACTORS)
RMD_MAP = dict(spec.RMD_DIVISORS)


# ---------------------------------------------------------------- life expectancy
def build_le():
    raw = [(D["cur_age_le"], D["le_cur"]), (62, D["le62"]), (67, D["le67"]), (70, D["le70"])]
    raw += [(a, D["le70"] * r) for a, r in spec.LE_TAIL_ANCHORS]
    raw_ages = [a for a, _ in raw]
    raw_les = [l for _, l in raw]
    ages_sorted = sorted(raw_ages)
    anchors = []
    for i, a in enumerate(ages_sorted):
        # INDEX(raw_les, MATCH(SMALL(raw_ages,k), raw_ages, 0)) - first exact match
        le_val = raw_les[raw_ages.index(a)]
        anchors.append((a + i * 0.0001, le_val))
    aa = [x for x, _ in anchors]
    al_ = [y for _, y in anchors]

    def match_le(age):
        pos = None
        for i, v in enumerate(aa, start=1):
            if v <= age:
                pos = i
            else:
                break
        return pos if pos else 1

    single = {}
    for age in AGES:
        m = match_le(age)
        lo_a, lo_l = aa[m - 1], al_[m - 1]
        j = min(m + 1, len(aa))
        hi_a, hi_l = aa[j - 1], al_[j - 1]
        inc = 0.0 if hi_a == lo_a else (age - lo_a) / (hi_a - lo_a) * (hi_l - lo_l)
        single[age] = max(round(lo_l + inc, 1), 0.1)

    p, S = {}, {}
    for age in AGES:
        nxt = single[age + 1] if age < spec.LE_LAST_AGE else single[age] * 0.84
        p[age] = min(max(single[age] / (1 + nxt), 0.0), 1.0)
    acc = 1.0
    for age in AGES:
        S[age] = acc
        acc *= p[age]
    gap = D["spouse_gap"]
    Sp = {}
    for age in AGES:
        sa = age - gap
        Sp[age] = S[sa] if sa in S else (1.0 if sa < spec.LE_FIRST_AGE else 0.0)
    joint = {}
    for age in AGES:
        if S[age] == 0 or Sp[age] == 0:
            joint[age] = single[age]
            continue
        joint[age] = sum(1 - (1 - S[a] / S[age]) * (1 - Sp[a] / Sp[age])
                         for a in AGES if a > age)
    planning = {age: max(joint[age] if D["household"] == "Couple" else single[age], 0.5)
                for age in AGES}
    return single, p, S, Sp, joint, planning


SINGLE, PSURV, SURV, SPSURV, JOINT, PLANNING = build_le()


def configure(overrides=None):
    """Reset to defaults, apply overrides, and rebuild the life-expectancy engine.

    Lets `test_scenarios.py` drive the shadow model with the same alternate inputs
    it pokes into the workbook.
    """
    global SINGLE, PSURV, SURV, SPSURV, JOINT, PLANNING
    D.clear()
    D.update(spec.DEFAULTS)
    if overrides:
        D.update(overrides)
    SINGLE, PSURV, SURV, SPSURV, JOINT, PLANNING = build_le()


def plan_years(age):
    return PLANNING.get(age, 1)


# ---------------------------------------------------------------- helpers
def ss_factor():
    return SS_MAP[D["ss_age"]]


def rmd_divisor(age):
    a = min(int(age), 120)
    if a in RMD_MAP:
        return RMD_MAP[a]
    best = None
    for k in sorted(RMD_MAP):
        if k <= a:
            best = k
    return RMD_MAP[best] if best else RMD_MAP[72]


def ord_rate(age):
    """Effective ordinary tax rate for the retirement phase this age falls in."""
    if age >= D["rmd_age"]:
        return D["tax_rmd"]
    if age >= D["ss_age"]:
        return D["tax_ss"]
    return D["tax_early"]


def source(w, tx, td):
    """Split a gross withdrawal taxable-first, then traditional, then Roth."""
    f_tx = min(w, tx)
    f_td = min(max(w - tx, 0.0), td)
    f_rt = max(w - f_tx - f_td, 0.0)
    return f_tx, f_td, f_rt


def gain_fraction(tx, basis):
    return max(1 - basis / tx, 0.0) if tx > 0 else 0.0


def rmd_from(age, td):
    if age >= D["rmd_age"]:
        return td / rmd_divisor(age)
    return 0.0


def step_accounts(age, tx, td, rt, basis, w, r):
    """Advance one year. Returns (after-tax spendable, next tx/td/rt/basis, tax pieces)."""
    o = ord_rate(age)
    f_tx, f_td, f_rt = source(w, tx, td)
    gain = gain_fraction(tx, basis)
    rmd_extra = max(min(rmd_from(age, td), td) - f_td, 0.0)
    tax_spend = f_tx * gain * D["cg_rate"] + f_td * o
    reinvest = rmd_extra * (1 - o)
    div = max(tx - f_tx, 0.0) * D["div_yield"]
    g = (1 + r) * (1 - D["fee"])
    ntx = max((tx - f_tx + reinvest) * g - div * D["cg_rate"], 0.0)
    ntd = max((td - f_td - rmd_extra) * g, 0.0)
    nrt = max((rt - f_rt) * g, 0.0)
    nbasis = max(basis * (1 - f_tx / tx if tx > 0 else 0.0) + reinvest + div, 0.0)
    return dict(f_tx=f_tx, f_td=f_td, f_rt=f_rt, gain=gain, rmd=rmd_from(age, td),
                rmd_extra=rmd_extra, tax_wd=f_tx * gain * D["cg_rate"]
                + (f_td + rmd_extra) * o, tax_spend=tax_spend,
                div_tax=div * D["cg_rate"], ord_rate=o,
                ntx=ntx, ntd=ntd, nrt=nrt, nbasis=nbasis)


def ceiling_for(age):
    """The guardrail ceiling that applies at this age.

    Fixed %       - the flat input, exactly as the source article specifies.
    Age-graduated - a multiple of the IRS Uniform Lifetime withdrawal rate, so the
                    cap rises with age instead of progressively overriding the
                    life-expectancy rule. The age is clamped to the published
                    table's 72-120 range.

    Ages outside the Life Expectancy tab's 35-110 range fall back to the fixed
    ceiling, mirroring the IFERROR on the workbook's VLOOKUP. Those rows sit past
    any legal planning horizon and never reach a result.
    """
    if age not in PLANNING:
        return D["ceiling"]
    if D["ceiling_basis"] == "Fixed %":
        return D["ceiling"]
    return D["ceiling_mult"] / rmd_divisor(age)


def need_today(year, age):
    """Total spending requirement for the year, in today's dollars, after tax."""
    n = D["essential"] * (1 + D["drift"]) ** (year - 1)
    if age < 65:
        n += D["healthcare"]
    if D["ltc_on"] == "Yes" and D["ltc_age"] <= age < D["ltc_age"] + D["ltc_years"]:
        n += D["ltc_cost"]
    return n


def exhausted_label(rows, col):
    """Mirror of the workbook's account-exhausted result rows.

    Excel uses MATCH(0, range, 0) — an exact match — which is safe because each
    ending balance is wrapped in MAX(..., 0) and therefore lands on a true zero.
    """
    age = next((r["B"] for r in rows if r[col] == 0), None)
    return "not within the plan" if age is None else f"age {age}"


def scen_return(year):
    s = D["scenario"]
    if s == "Steady":
        return D["exp_ret"]
    if s == "Early crash":
        if year <= spec.SCENARIO_YEARS:
            return spec.CRASH_RETURNS.get(year, D["exp_ret"])
        return D["exp_ret"]
    if year <= spec.SCENARIO_YEARS:
        return spec.STAGFLATION_RETURN if year <= spec.STAGFLATION_YEARS else D["exp_ret"]
    return D["exp_ret"]


def scen_infl(year):
    s = D["scenario"]
    if s == "Steady":
        return D["infl"]
    if s == "Early crash":
        return D["infl"]
    if year <= spec.SCENARIO_YEARS:
        return spec.STAGFLATION_INFLATION if year <= spec.STAGFLATION_YEARS else D["infl"]
    return D["infl"]


# ---------------------------------------------------------------- 4% rule
def four_rule():
    rows = []
    for k in range(N):
        yr = k + 1
        age = D["start_age"] + k
        c = scen_return(yr)
        d = scen_infl(yr)
        if k == 0:
            e = 1.0
            g_ = D["taxable0"]
            h = D["traditional0"]
            i = D["roth0"]
            j = D["taxable0"] * D["basis_share"]
            kk = g_ + h + i
            l = D["current_4pct"] if D["current_4pct"] > 0 else kk * D["rate4"]
        else:
            pr = rows[-1]
            e = pr["E"] * (1 + pr["D"])
            g_ = max(pr["AA"], 0.0)
            h = max(pr["AB"], 0.0)
            i = max(pr["AC"], 0.0)
            j = max(pr["AD"], 0.0)
            kk = g_ + h + i
            l = pr["L"] * (1 + pr["D"])
        f = ord_rate(age)
        m = min(l, kk)
        st = step_accounts(age, g_, h, i, j, m, c)
        n, o, p_ = st["f_tx"], st["f_td"], st["f_rt"]
        qq = st["rmd"]
        r = st["rmd_extra"]
        s = st["tax_wd"]
        t = st["div_tax"]
        u = D["full_ss"] * ss_factor() * e if age >= D["ss_age"] else 0.0
        v = u * spec.SS_TAXABLE_SHARE * f
        w = m - st["tax_spend"] + u - v
        x = w / e
        y = need_today(yr, age)
        z = x - y
        rows.append(dict(A=yr, B=age, C=c, D=d, E=e, F=f, G=g_, H=h, I=i, J=j, K=kk, L=l,
                         M=m, N=n, O=o, P=p_, Q=qq, R=r, S=s, T=t, U=u, V=v, W=w, X=x,
                         Y=y, Z=z, AA=st["ntx"], AB=st["ntd"], AC=st["nrt"], AD=st["nbasis"],
                         AE=st["ntx"] + st["ntd"] + st["nrt"],
                         AF=(st["ntx"] + st["ntd"] + st["nrt"]) / (e * (1 + d))))
    return rows


# ---------------------------------------------------------------- dynamic
def dynamic(shared):
    rows = []
    for k in range(N):
        yr = k + 1
        age = D["start_age"] + k
        c = plan_years(age)
        d = shared[k]["C"]
        e = shared[k]["D"]
        f = shared[k]["E"]
        g_ = ord_rate(age)
        if k == 0:
            h = D["taxable0"]
            i = D["traditional0"]
            j = D["roth0"]
            kk = D["taxable0"] * D["basis_share"]
        else:
            pr = rows[-1]
            h = max(pr["AE"], 0.0)
            i = max(pr["AF"], 0.0)
            j = max(pr["AG"], 0.0)
            kk = max(pr["AH"], 0.0)
        l = h + i + j
        m = l / max(c, 0.5)
        n = min(ceiling_for(age) * l, max(D["floor"] * l, m))
        if k == 0:
            o = n
        else:
            pr = rows[-1]
            stp = (1 + pr["E"]) if D["shock_basis"] == "Real" else 1.0
            prev = pr["P"] * stp
            o = min(prev * (1 + D["shock"]), max(prev * (1 - D["shock"]), n))
        p_ = min(o, l)
        qq = p_ / l if l else 0.0
        st = step_accounts(age, h, i, j, kk, p_, d)
        r, s, t = st["f_tx"], st["f_td"], st["f_rt"]
        u = st["rmd"]
        v = st["rmd_extra"]
        w = st["tax_wd"]
        x = st["div_tax"]
        y = D["full_ss"] * ss_factor() * f if age >= D["ss_age"] else 0.0
        z = y * spec.SS_TAXABLE_SHARE * g_
        aa = p_ - st["tax_spend"] + y - z
        ab = aa / f
        ac = need_today(yr, age)
        ad = ab - ac
        tot = st["ntx"] + st["ntd"] + st["nrt"]
        rows.append(dict(A=yr, B=age, C=c, D=d, E=e, F=f, G=g_, H=h, I=i, J=j, K=kk, L=l,
                         M=m, N=n, O=o, P=p_, Q=qq, R=r, S=s, T=t, U=u, V=v, W=w, X=x,
                         Y=y, Z=z, AA=aa, AB=ab, AC=ac, AD=ad,
                         AE=st["ntx"], AF=st["ntd"], AG=st["nrt"], AH=st["nbasis"],
                         AI=tot, AJ=tot / (f * (1 + e)), AK=ceiling_for(age)))
    return rows



# ---------------------------------------------------------------- monte carlo
def monte_carlo():
    rng = random.Random(spec.MC_SEED)
    draws = []
    for _ in range(spec.MAX_SIMS):
        ltc_u = rng.random()
        yrs = [(rng.gauss(0, 1), rng.gauss(0, 1), rng.randrange(1, len(spec.HISTORICAL) + 1))
               for _ in range(N)]
        draws.append((ltc_u, yrs))
    pos = min(max(D["plan_age"] - D["start_age"], 1), N)
    parametric = D["mc_method"] == "Parametric"
    out = []
    for sim in range(spec.MAX_SIMS):
        ltc_u, yrs = draws[sim]
        f_tx, f_td, f_rt, f_bs = (D["taxable0"], D["traditional0"], D["roth0"],
                                  D["taxable0"] * D["basis_share"])
        d_tx, d_td, d_rt, d_bs = f_tx, f_td, f_rt, f_bs
        f_target = (D["current_4pct"] if D["current_4pct"] > 0
                    else (f_tx + f_td + f_rt) * D["rate4"])
        d_prev = None
        index = 1.0
        f_sp, d_sp, needs, changes = [], [], [], []
        f_total = d_total = 0.0
        last_infl = D["infl"]
        for y in range(1, pos + 1):
            z_ret, z_inf, hrow = yrs[y - 1]
            age = D["start_age"] + y - 1
            if parametric:
                ret = math.exp(math.log(1 + D["exp_ret"]) - 0.5 * D["mc_vol"] ** 2
                               + D["mc_vol"] * z_ret) - 1
                inf = max(D["infl"] + D["mc_infl_vol"] * z_inf, spec.INFLATION_FLOOR)
            else:
                _, s_, b_, cpi_ = spec.HISTORICAL[hrow - 1]
                ret = D["mc_stock"] * s_ + (1 - D["mc_stock"]) * b_
                inf = cpi_
            pyrs = plan_years(age)
            o = ord_rate(age)
            ss = D["full_ss"] * ss_factor() * index if age >= D["ss_age"] else 0.0
            # 4% rule
            f_total = f_tx + f_td + f_rt
            w4 = min(f_target, f_total)
            st4 = step_accounts(age, f_tx, f_td, f_rt, f_bs, w4, ret)
            f_sp.append((w4 - st4["tax_spend"]
                         + ss * (1 - spec.SS_TAXABLE_SHARE * o)) / index)
            # dynamic
            d_total = d_tx + d_td + d_rt
            guarded = min(ceiling_for(age) * d_total,
                          max(D["floor"] * d_total, d_total / max(pyrs, 0.5)))
            if d_prev is None:
                wd = min(guarded, d_total)
                changes.append(None)
            else:
                stp = (1 + last_infl) if D["shock_basis"] == "Real" else 1.0
                prev = d_prev * stp
                wd = min(min(prev * (1 + D["shock"]),
                             max(prev * (1 - D["shock"]), guarded)), d_total)
                changes.append(wd / d_prev - 1 if d_prev > 0 else 0.0)
            std = step_accounts(age, d_tx, d_td, d_rt, d_bs, wd, ret)
            d_sp.append((wd - std["tax_spend"]
                         + ss * (1 - spec.SS_TAXABLE_SHARE * o)) / index)
            nd = D["essential"] * (1 + D["drift"]) ** (y - 1)
            if age < 65:
                nd += D["healthcare"]
            if (D["ltc_on"] == "Yes" and ltc_u < D["mc_ltc_prob"]
                    and D["ltc_age"] <= age < D["ltc_age"] + D["ltc_years"]):
                nd += D["ltc_cost"]
            needs.append(nd)
            f_tx, f_td, f_rt, f_bs = st4["ntx"], st4["ntd"], st4["nrt"], st4["nbasis"]
            d_tx, d_td, d_rt, d_bs = std["ntx"], std["ntd"], std["nrt"], std["nbasis"]
            d_prev = wd
            f_target = f_target * (1 + inf)
            last_infl = inf
            index = index * (1 + inf)
        # the balance "at plan-to age" is the opening balance of the following year
        f_end = f_tx + f_td + f_rt
        d_end = d_tx + d_td + d_rt
        deflator = index
        real_changes = [c for c in changes[1:] if c is not None]
        out.append(dict(
            four_lasts=1 if f_end > 0 else 0,
            four_ess=1 if all(a >= b for a, b in zip(f_sp, needs)) else 0,
            four_bal=f_end / deflator,
            four_low=min(f_sp), four_sum=sum(f_sp),
            dyn_lasts=1 if d_end > 0 else 0,
            dyn_ess=1 if all(a >= b for a, b in zip(d_sp, needs)) else 0,
            dyn_bal=d_end / deflator,
            dyn_low=min(d_sp), dyn_sum=sum(d_sp),
            four_share=sum(1 for a, b in zip(f_sp, needs) if a >= b) / pos,
            dyn_share=sum(1 for a, b in zip(d_sp, needs) if a >= b) / pos,
            dyn_worst=min(real_changes) if real_changes else 0.0))
    return out


# ---------------------------------------------------------------- comparison
class Checker:
    def __init__(self):
        self.checked = 0
        self.fails = []

    def eq(self, where, expected, actual, tol=1e-6):
        self.checked += 1
        if expected is None or actual is None:
            if expected != actual:
                self.fails.append((where, expected, actual))
            return
        if isinstance(expected, str) or isinstance(actual, str):
            if str(expected).strip() != str(actual).strip():
                self.fails.append((where, expected, actual))
            return
        scale = max(1.0, abs(expected), abs(actual))
        if abs(expected - actual) > tol * scale:
            self.fails.append((where, expected, actual))


def main():
    if not os.path.exists(BOOK):
        print("MISSING WORKBOOK - run build_retirement_planner.py then recalc_excel.ps1")
        return 1
    wb = load_workbook(BOOK, data_only=True)
    ck = Checker()

    shared = four_rule()
    dyn = dynamic(shared)

    # ---- Life Expectancy tab ----
    le = wb["Life Expectancy"]
    for i, age in enumerate(AGES):
        r = 6 + i
        ck.eq(f"LE!A{r}", age, le.cell(r, 1).value)
        ck.eq(f"LE!B{r}", SINGLE[age], le.cell(r, 2).value, 1e-9)
        ck.eq(f"LE!H{r}", PSURV[age], le.cell(r, 8).value, 1e-9)
        ck.eq(f"LE!I{r}", SURV[age], le.cell(r, 9).value, 1e-9)
        ck.eq(f"LE!J{r}", SPSURV[age], le.cell(r, 10).value, 1e-9)
        ck.eq(f"LE!K{r}", JOINT[age], le.cell(r, 11).value, 1e-7)
        ck.eq(f"LE!L{r}", PLANNING[age], le.cell(r, 12).value, 1e-9)
        ck.eq(f"LE!M{r}", ceiling_for(age), le.cell(r, 13).value, 1e-9)

    # ---- 4% Rule tab ----
    fr = wb["4% Rule"]
    cols4 = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
             "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC",
             "AD", "AE", "AF"]
    for k, row in enumerate(shared):
        r = R0 + k
        for ci, letter in enumerate(cols4, start=1):
            ck.eq(f"4%!{letter}{r}", row[letter], fr.cell(r, ci).value)

    # ---- Dynamic tab ----
    dysh = wb["Dynamic Strategy"]
    colsd = cols4 + ["AG", "AH", "AI", "AJ", "AK"]
    for k, row in enumerate(dyn):
        r = R0 + k
        for ci, letter in enumerate(colsd, start=1):
            ck.eq(f"DYN!{letter}{r}", row[letter], dysh.cell(r, ci).value)

    # ---- Monte Carlo outcomes ----
    mc_out = monte_carlo()
    oc = wb["MC Outcomes"]
    keys = [(2, "four_lasts"), (3, "four_ess"), (4, "four_bal"), (5, "four_low"),
            (6, "four_sum"), (7, "dyn_lasts"), (8, "dyn_ess"), (9, "dyn_bal"),
            (10, "dyn_low"), (11, "dyn_sum"), (12, "dyn_worst"),
            (13, "four_share"), (14, "dyn_share")]
    for sim, exp in enumerate(mc_out):
        r = 5 + sim
        for col, key in keys:
            ck.eq(f"MCOUT!{chr(64 + col)}{r}", exp[key], oc.cell(r, col).value, 1e-6)

    # ---- headline results (section-aware: both blocks reuse the same labels) ----
    horizon = min(max(D["plan_age"] - D["start_age"], 1), N)
    inputs = wb["Inputs & Summary"]
    section = None
    found = {}
    for r in range(1, 90):
        lab = inputs.cell(r, 5).value
        if not isinstance(lab, str):
            continue
        lab = lab.strip()
        if lab.startswith("—") and lab.endswith("—"):
            section = lab.strip("— ").strip()
            continue
        found[(section, lab)] = inputs.cell(r, 6).value

    dyn_h = dyn[:horizon]
    four_h = shared[:horizon]
    expectations = [
        (None, "Years modelled (start age to plan-to age)",
         D["plan_age"] - D["start_age"]),
        (None, "Net expected return after fees",
         (1 + D["exp_ret"]) * (1 - D["fee"]) - 1),
        ("DOES THE PLAN WORK?", "Dynamic: essentials covered every year?",
         "YES" if all(r["AD"] >= 0 for r in dyn_h) else "NO"),
        ("DOES THE PLAN WORK?", "4% rule: essentials covered every year?",
         "YES" if all(r["Z"] >= 0 for r in four_h) else "NO"),
        ("DYNAMIC STRATEGY", "First-year after-tax income (today's $)", dyn[0]["AB"]),
        ("DYNAMIC STRATEGY", "Lowest annual after-tax income (today's $)",
         min(r["AB"] for r in dyn_h)),
        ("DYNAMIC STRATEGY", "Highest annual after-tax income (today's $)",
         max(r["AB"] for r in dyn_h)),
        ("DYNAMIC STRATEGY", "Lifetime after-tax income (today's $)",
         sum(r["AB"] for r in dyn_h)),
        ("DYNAMIC STRATEGY", "Years spending falls short of needs",
         sum(1 for r in dyn_h if r["AD"] < 0)),
        ("DYNAMIC STRATEGY", "Share of years essentials are covered",
         1 - sum(1 for r in dyn_h if r["AD"] < 0) / horizon),
        ("DYNAMIC STRATEGY", "Real income at plan-to age vs. year 1",
         dyn[horizon - 1]["AB"] / dyn[0]["AB"] - 1),
        ("DYNAMIC STRATEGY", "Balance at plan-to age (today's $)", dyn[horizon - 1]["AJ"]),
        ("DYNAMIC STRATEGY", "Money lasts to plan-to age?",
         "YES" if dyn[horizon - 1]["AI"] > 0 else "NO"),
        ("DYNAMIC STRATEGY", "Guardrail ceiling starts capping spending at",
         (lambda a: "never — the rule runs freely" if a is None
          else f"age {a} (of {horizon} years modelled)")(
             next((r["B"] for r in dyn_h
                   if 1 / r["C"] > ceiling_for(r["B"]) + 1e-6), None))),
        ("DYNAMIC STRATEGY", "Brokerage accounts exhausted at",
         exhausted_label(dyn_h, "AE")),
        ("DYNAMIC STRATEGY", "Traditional IRA accounts exhausted at",
         exhausted_label(dyn_h, "AF")),
        ("DYNAMIC STRATEGY", "Roth accounts exhausted at",
         exhausted_label(dyn_h, "AG")),
        ("4% RULE", "First-year after-tax income (today's $)", shared[0]["X"]),
        ("4% RULE", "Lowest annual after-tax income (today's $)",
         min(r["X"] for r in four_h)),
        ("4% RULE", "Lifetime after-tax income (today's $)",
         sum(r["X"] for r in four_h)),
        ("4% RULE", "Years spending falls short of needs",
         sum(1 for r in four_h if r["Z"] < 0)),
        ("4% RULE", "Share of years essentials are covered",
         1 - sum(1 for r in four_h if r["Z"] < 0) / horizon),
        ("4% RULE", "Balance at plan-to age (today's $)", shared[horizon - 1]["AF"]),
        ("4% RULE", "Money lasts to plan-to age?",
         "YES" if shared[horizon - 1]["AE"] > 0 else "NO"),
        ("WHAT YOUR ESSENTIALS COST", "Social Security you will actually receive "
         "(today's $/yr, after tax)",
         D["full_ss"] * ss_factor() * (1 - spec.SS_TAXABLE_SHARE * D["tax_ss"])),
    ]
    for sect, label, expected in expectations:
        key = (sect, label)
        if key not in found:
            ck.fails.append((f"RESULT[{sect} / {label}]", expected, "<label not found>"))
            ck.checked += 1
            continue
        ck.eq(f"RESULT[{sect} / {label}]", expected, found[key], 1e-6)

    print(f"cells compared: {ck.checked:,}")
    if ck.fails:
        print(f"STATUS: MISMATCHES - {len(ck.fails)}")
        for where, exp, act in ck.fails[:40]:
            print(f"  {where}: expected {exp!r}  got {act!r}")
        if len(ck.fails) > 40:
            print(f"  ... and {len(ck.fails) - 40} more")
        return 1
    print("STATUS: success - model matches Excel on every compared cell")
    return 0


if __name__ == "__main__":
    sys.exit(main())
